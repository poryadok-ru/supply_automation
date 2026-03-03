# import pandas as pd
# from decimal import Decimal, ROUND_HALF_UP
# import numpy as np
# import glob
# from time import time
# from datetime import datetime
# import os
# import shutil
# from python_calamine.pandas import pandas_monkeypatch
# from openpyxl import Workbook
# from openpyxl.styles import PatternFill
# from openpyxl.utils.dataframe import dataframe_to_rows


# pandas_monkeypatch()

# pd.options.mode.chained_assignment = None


# os.chdir(os.path.dirname(os.path.abspath(__file__)))

# def nacenkaview(file_path, add_message):
#     if not file_path.endswith(os.sep):
#         file_path = file_path + os.sep

#     t = time()
#     add_message('Стартуем!')


#     def custom_round_decimal(number):
#         decimal_number = Decimal(str(number))
#         rounded_number = decimal_number.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
#         return float(rounded_number)
    
#     d = {f: os.stat(f).st_mtime for f in glob.iglob(file_path + "Исходник\\*.xlsx")}
#     file_name_otchet = max(d, key=lambda i: d[i])

#     df = pd.read_excel(file_name_otchet, engine='calamine', skiprows=7)

#     df_rrc = pd.read_excel(r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Шафиев\Наценка\РРЦ\РРЦ.xlsx",
#                         engine='calamine', usecols=['Код', 'Розничная_ИЦ', 'Интернет-магазина', 'Оптовая_1'])

#     df_rrc = df_rrc.rename(columns={'Розничная_ИЦ': 'Розничная_ИЦРРЦ',
#                                     'Интернет-магазина': 'МП Poryadok.ru до скидкиРРЦ',
#                                     'Оптовая_1': 'Оптовая_1РРЦ'})

#     index_zakup = df.columns.get_loc('Закупочная')
#     index_mp_poryadok = df.columns.get_loc('МП Poryadok.ru до скидки')
#     index_opt = df.columns.get_loc('Оптовая 1')
#     index_rozn = df.columns.get_loc('Розничная')

#     df = df.rename(columns={'Закупочная': 'Закупочная1',
#                             'Unnamed: ' + str(index_zakup + 1): 'Закупочная2',
#                             'Unnamed: ' + str(index_zakup + 2): 'ЗакупочнаяИЗМ',
#                             'МП Poryadok.ru до скидки': 'МП Poryadok.ru до скидки_ИЦ1',
#                             'Unnamed: ' + str(index_mp_poryadok + 3): 'МП Poryadok.ru до скидки_ИЦ_Последняя цена за период',
#                             'Оптовая 1': 'Оптовая 1_ИЦ1',
#                             'Unnamed: ' + str(index_opt + 3): 'Оптовая 1_Последняя цена за период',
#                             'Розничная': 'Розничная_ИЦ1',
#                             'Unnamed: ' + str(index_rozn + 3): 'Розничная_ИЦ_Последняя цена за период'})


#     df = df[1:].reset_index(drop=True)
#     df = df[['Код', 'Номенклатура', 'Закупочная1', 'Закупочная2', 'ЗакупочнаяИЗМ',
#             'МП Poryadok.ru до скидки_ИЦ1', 'МП Poryadok.ru до скидки_ИЦ_Последняя цена за период', 'Оптовая 1_ИЦ1',
#             'Оптовая 1_Последняя цена за период', 'Розничная_ИЦ1',
#             'Розничная_ИЦ_Последняя цена за период']]



#     df['Код'] = pd.to_numeric(df['Код'], errors='coerce')
#     df['Розничная_ИЦ2'] = df['Розничная_ИЦ_Последняя цена за период']
#     df['Оптовая 1_ИЦ2'] = df['Оптовая 1_Последняя цена за период']
#     df['МП Poryadok.ru до скидки_ИЦ2'] = df['МП Poryadok.ru до скидки_ИЦ_Последняя цена за период']

#     # Преобразование столбцов в числовой формат
#     numeric_columns = ['Закупочная1', 'Закупочная2', 'ЗакупочнаяИЗМ',
#                     'МП Poryadok.ru до скидки_ИЦ1', 'МП Poryadok.ru до скидки_ИЦ_Последняя цена за период', 'Оптовая 1_ИЦ1',
#                     'Оптовая 1_Последняя цена за период', 'Розничная_ИЦ1', 'Розничная_ИЦ_Последняя цена за период', 'Розничная_ИЦ_Последняя цена за период', 
#                     'Оптовая 1_Последняя цена за период', 'МП Poryadok.ru до скидки_ИЦ_Последняя цена за период']

#     # Преобразование данных
#     for col in numeric_columns:
#         # Приведение всех значений к строковому типу перед использованием .str.replace
#         df[col] = df[col].astype(str).str.replace(',', '.').astype(float)


#     df = df.merge(df_rrc, on='Код', how='left')
#     # df.to_excel("1234.xlsx", index=False)

#     df.loc[(~df['Розничная_ИЦРРЦ'].isna()) & (~df['Розничная_ИЦ_Последняя цена за период'].isna()), 'Розничная_ИЦ_Последняя цена за период'] = df['Розничная_ИЦРРЦ']
#     df.loc[(~df['Розничная_ИЦРРЦ'].isna()) & (~df['ЗакупочнаяИЗМ'].isna()), 'Розничная_ИЦ_Последняя цена за период'] = df['Розничная_ИЦРРЦ']
#     df.loc[(~df['Розничная_ИЦРРЦ'].isna()) & (df['Розничная_ИЦ1'] < df['Розничная_ИЦРРЦ']), 'Розничная_ИЦ_Последняя цена за период'] = df['Розничная_ИЦРРЦ']
#     df.loc[(~df['Розничная_ИЦРРЦ'].isna()) & (df['Розничная_ИЦ2'] > df['Розничная_ИЦРРЦ']), 'Розничная_ИЦ_Последняя цена за период'] = df['Розничная_ИЦ2']
#     df.loc[(~df['Оптовая_1РРЦ'].isna()) & (~df['Оптовая 1_Последняя цена за период'].isna()), 'Оптовая 1_Последняя цена за период'] = df['Оптовая_1РРЦ']
#     df.loc[(~df['Оптовая_1РРЦ'].isna()) & (~df['ЗакупочнаяИЗМ'].isna()), 'Оптовая 1_Последняя цена за период'] = df['Оптовая_1РРЦ']
#     df.loc[(~df['Оптовая_1РРЦ'].isna()) & (df['Оптовая 1_ИЦ1'] < df['Оптовая_1РРЦ']), 'Оптовая 1_Последняя цена за период'] = df['Оптовая_1РРЦ']
#     df.loc[(~df['Оптовая_1РРЦ'].isna()) & (df['Оптовая 1_ИЦ2'] > df['Оптовая_1РРЦ']), 'Оптовая 1_Последняя цена за период'] = df['Оптовая 1_ИЦ2']
#     df.loc[(~df['МП Poryadok.ru до скидкиРРЦ'].isna()) &
#         (~df['МП Poryadok.ru до скидки_ИЦ_Последняя цена за период'].isna()), 'МП Poryadok.ru до скидки_ИЦ_Последняя цена за период'] = df['МП Poryadok.ru до скидкиРРЦ']
#     df.loc[(~df['МП Poryadok.ru до скидкиРРЦ'].isna()) & (~df['ЗакупочнаяИЗМ'].isna()), 'МП Poryadok.ru до скидки_ИЦ_Последняя цена за период'] = df['МП Poryadok.ru до скидкиРРЦ']
#     df.loc[(~df['МП Poryadok.ru до скидкиРРЦ'].isna()) & (df['МП Poryadok.ru до скидки_ИЦ1'] < df['МП Poryadok.ru до скидкиРРЦ']), 'МП Poryadok.ru до скидки_ИЦ_Последняя цена за период'] = df['МП Poryadok.ru до скидкиРРЦ']
#     df.loc[(~df['МП Poryadok.ru до скидкиРРЦ'].isna()) & (df['МП Poryadok.ru до скидки_ИЦ2'] > df['МП Poryadok.ru до скидкиРРЦ']), 'МП Poryadok.ru до скидки_ИЦ_Последняя цена за период'] = df['МП Poryadok.ru до скидки_ИЦ2']
#     # df.to_excel("12345.xlsx", index=False)

    
#     # Рассчитываем "Оптовая 1" только для строк, где "изменение" является числом
#     df['Оптовая 1'] = np.where(
#         (~df['Оптовая 1_Последняя цена за период'].isna()) & (~df['Закупочная2'].isna()),  # Проверяем, что "изменение" и "Закупочная2" не NaN
#         (df['Оптовая 1_Последняя цена за период'] / df['Закупочная2'] * 100) - 100,   # Выполняем расчёт
#         np.nan)

#     df['Розничная'] = np.where(
#         (~df['Розничная_ИЦ_Последняя цена за период'].isna()) & (~df['Закупочная2'].isna()),  # Проверяем, что "изменение" и "Закупочная2" не NaN
#         (df['Розничная_ИЦ_Последняя цена за период'] / df['Закупочная2'] * 100) - 100,   # Выполняем расчёт
#         np.nan)

#     df['МП Poryadok.ru до скидки'] = np.where(
#         (~df['МП Poryadok.ru до скидки_ИЦ_Последняя цена за период'].isna()) & (~df['Закупочная2'].isna()),  # Проверяем, что "изменение" и "Закупочная2" не NaN
#         (df['МП Poryadok.ru до скидки_ИЦ_Последняя цена за период'] / df['Закупочная2'] * 100) - 100,   # Выполняем расчёт
#         np.nan)

    
#     df['Розничная'] = (df['Розничная']).apply(custom_round_decimal)
#     df['Оптовая 1'] = (df['Оптовая 1']).apply(custom_round_decimal)
#     df['МП Poryadok.ru до скидки'] = (df['МП Poryadok.ru до скидки']).apply(custom_round_decimal)

#     # Добавляем столбец с отметкой о наличии РРЦ
#     df['Есть РРЦ'] = df['Оптовая_1РРЦ'].notna()
    


#     df1 = df[~df['Оптовая 1'].isna()][['Код', 'Номенклатура', 'Оптовая 1']]
#     df2 = df[~df['Розничная'].isna()][['Код', 'Номенклатура', 'Розничная']]
#     df3 = df[~df['МП Poryadok.ru до скидки'].isna()][['Код', 'Номенклатура', 'МП Poryadok.ru до скидки']]


#     now1 = datetime.now()
#     now1 = now1.strftime("%d-%m-%Y")

    

#     df1.to_excel(file_path + 'Для загрузки Оптовая 1' + ' ' + now1 +'.xlsx', index=False)

#     # Создаем новую книгу
#     wb = Workbook()
#     ws = wb.active

#     # Записываем данные из DataFrame
#     for r_idx, row in enumerate(dataframe_to_rows(df1, index=False, header=True), 1):
#         for c_idx, value in enumerate(row, 1):
#             ws.cell(row=r_idx, column=c_idx, value=value)

#     # Применяем красную заливку для строк с РРЦ
#     red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')

#     for row_idx in range(2, ws.max_row + 1):  # Пропускаем заголовок
#         has_rrc = ws.cell(row=row_idx, column=4).value  # Столбец 'Есть РРЦ'
#         if has_rrc:
#             for col_idx in range(1, ws.max_column + 1):
#                 ws.cell(row=row_idx, column=col_idx).fill = red_fill

#     # Удаляем столбец 'Есть РРЦ' если нужно (опционально)
#     ws.delete_cols(4)

#     # Сохраняем файл
#     wb.save(file_path + 'Для загрузки Оптовая 1' + ' ' + now1 +'.xlsx')
#     df2.to_excel(file_path + 'Для загрузки Розничная' + ' ' + now1 +'.xlsx', index=False)
#     df3.to_excel(file_path + 'Для загрузки МП Poryadok.ru до скидки' + ' ' + now1 +'.xlsx', index=False)

#     destination_folder = file_path + r'архив_отчетов'
#     if not os.path.exists(destination_folder):
#         os.makedirs(destination_folder)

    
#     date_folder = os.path.join(destination_folder, now1)
#     # Создание папки с текущей датой
#     if not os.path.exists(date_folder):
#         os.makedirs(date_folder)
#     shutil.copy2(file_name_otchet, os.path.join(date_folder))
    

#     add_message(f'Готово за {time() - t}')


import pandas as pd
from decimal import Decimal, ROUND_HALF_UP
import numpy as np
import glob
from time import time
from datetime import datetime
import os
import shutil
from python_calamine.pandas import pandas_monkeypatch
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


pandas_monkeypatch()

pd.options.mode.chained_assignment = None


os.chdir(os.path.dirname(os.path.abspath(__file__)))

def nacenkaview(file_path, add_message):
    if not file_path.endswith(os.sep):
        file_path = file_path + os.sep

    t = time()
    add_message('Стартуем!')


    def custom_round_decimal(number):
        decimal_number = Decimal(str(number))
        rounded_number = decimal_number.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
        return float(rounded_number)
    
    d = {f: os.stat(f).st_mtime for f in glob.iglob(file_path + "Исходник\\*.xlsx")}
    file_name_otchet = max(d, key=lambda i: d[i])

    df = pd.read_excel(file_name_otchet, engine='calamine', skiprows=5)

    df_rrc = pd.read_excel(r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Шафиев\Наценка\РРЦ\РРЦ.xlsx",
                        engine='calamine', usecols=['Код', 'Розничная_ИЦ', 'Интернет-магазина', 'Оптовая_1'])

    df_rrc = df_rrc.rename(columns={'Розничная_ИЦ': 'Розничная_ИЦРРЦ',
                                    'Интернет-магазина': 'МП Poryadok.ru до скидкиРРЦ',
                                    'Оптовая_1': 'Оптовая_1РРЦ'})

    index_zakup = df.columns.get_loc('Закупочная')
    index_mp_poryadok = df.columns.get_loc('МП Poryadok.ru до скидки')
    index_opt = df.columns.get_loc('Оптовая 1')
    index_rozn = df.columns.get_loc('Розничная')

    df = df.rename(columns={'Закупочная': 'Закупочная1',
                            'Unnamed: ' + str(index_zakup + 1): 'Закупочная2',
                            'Unnamed: ' + str(index_zakup + 2): 'ЗакупочнаяИЗМ',
                            'МП Poryadok.ru до скидки': 'МП Poryadok.ru до скидки_ИЦ1',
                            'Unnamed: ' + str(index_mp_poryadok + 3): 'МП Poryadok.ru до скидки_ИЦ_Последняя цена за период',
                            'Оптовая 1': 'Оптовая 1_ИЦ1',
                            'Unnamed: ' + str(index_opt + 3): 'Оптовая 1_Последняя цена за период',
                            'Розничная': 'Розничная_ИЦ1',
                            'Unnamed: ' + str(index_rozn + 3): 'Розничная_ИЦ_Последняя цена за период'})


    df = df[1:].reset_index(drop=True)
    df = df[['Код', 'Номенклатура', 'Закупочная1', 'Закупочная2', 'ЗакупочнаяИЗМ',
            'МП Poryadok.ru до скидки_ИЦ1', 'МП Poryadok.ru до скидки_ИЦ_Последняя цена за период', 'Оптовая 1_ИЦ1',
            'Оптовая 1_Последняя цена за период', 'Розничная_ИЦ1',
            'Розничная_ИЦ_Последняя цена за период']]



    df['Код'] = pd.to_numeric(df['Код'], errors='coerce')
    df['Розничная_ИЦ2'] = df['Розничная_ИЦ_Последняя цена за период']
    df['Оптовая 1_ИЦ2'] = df['Оптовая 1_Последняя цена за период']
    df['МП Poryadok.ru до скидки_ИЦ2'] = df['МП Poryadok.ru до скидки_ИЦ_Последняя цена за период']

    # Преобразование столбцов в числовой формат
    numeric_columns = ['Закупочная1', 'Закупочная2', 'ЗакупочнаяИЗМ',
                    'МП Poryadok.ru до скидки_ИЦ1', 'МП Poryadok.ru до скидки_ИЦ_Последняя цена за период', 'Оптовая 1_ИЦ1',
                    'Оптовая 1_Последняя цена за период', 'Розничная_ИЦ1', 'Розничная_ИЦ_Последняя цена за период', 'Розничная_ИЦ_Последняя цена за период', 
                    'Оптовая 1_Последняя цена за период', 'МП Poryadok.ru до скидки_ИЦ_Последняя цена за период']

    # Преобразование данных
    for col in numeric_columns:
        # Приведение всех значений к строковому типу перед использованием .str.replace
        df[col] = df[col].astype(str).str.replace(',', '.').astype(float)


    df = df.merge(df_rrc, on='Код', how='left')
    # df.to_excel("1234.xlsx", index=False)

    df.loc[(~df['Розничная_ИЦРРЦ'].isna()) & (~df['Розничная_ИЦ_Последняя цена за период'].isna()), 'Розничная_ИЦ_Последняя цена за период'] = df['Розничная_ИЦРРЦ']
    df.loc[(~df['Розничная_ИЦРРЦ'].isna()) & (~df['ЗакупочнаяИЗМ'].isna()), 'Розничная_ИЦ_Последняя цена за период'] = df['Розничная_ИЦРРЦ']
    df.loc[(~df['Розничная_ИЦРРЦ'].isna()) & (df['Розничная_ИЦ1'] < df['Розничная_ИЦРРЦ']), 'Розничная_ИЦ_Последняя цена за период'] = df['Розничная_ИЦРРЦ']
    df.loc[(~df['Розничная_ИЦРРЦ'].isna()) & (df['Розничная_ИЦ2'] > df['Розничная_ИЦРРЦ']), 'Розничная_ИЦ_Последняя цена за период'] = df['Розничная_ИЦ2']
    df.loc[(~df['Оптовая_1РРЦ'].isna()) & (~df['Оптовая 1_Последняя цена за период'].isna()), 'Оптовая 1_Последняя цена за период'] = df['Оптовая_1РРЦ']
    df.loc[(~df['Оптовая_1РРЦ'].isna()) & (~df['ЗакупочнаяИЗМ'].isna()), 'Оптовая 1_Последняя цена за период'] = df['Оптовая_1РРЦ']
    df.loc[(~df['Оптовая_1РРЦ'].isna()) & (df['Оптовая 1_ИЦ1'] < df['Оптовая_1РРЦ']), 'Оптовая 1_Последняя цена за период'] = df['Оптовая_1РРЦ']
    df.loc[(~df['Оптовая_1РРЦ'].isna()) & (df['Оптовая 1_ИЦ2'] > df['Оптовая_1РРЦ']), 'Оптовая 1_Последняя цена за период'] = df['Оптовая 1_ИЦ2']
    df.loc[(~df['МП Poryadok.ru до скидкиРРЦ'].isna()) &
        (~df['МП Poryadok.ru до скидки_ИЦ_Последняя цена за период'].isna()), 'МП Poryadok.ru до скидки_ИЦ_Последняя цена за период'] = df['МП Poryadok.ru до скидкиРРЦ']
    df.loc[(~df['МП Poryadok.ru до скидкиРРЦ'].isna()) & (~df['ЗакупочнаяИЗМ'].isna()), 'МП Poryadok.ru до скидки_ИЦ_Последняя цена за период'] = df['МП Poryadok.ru до скидкиРРЦ']
    df.loc[(~df['МП Poryadok.ru до скидкиРРЦ'].isna()) & (df['МП Poryadok.ru до скидки_ИЦ1'] < df['МП Poryadok.ru до скидкиРРЦ']), 'МП Poryadok.ru до скидки_ИЦ_Последняя цена за период'] = df['МП Poryadok.ru до скидкиРРЦ']
    df.loc[(~df['МП Poryadok.ru до скидкиРРЦ'].isna()) & (df['МП Poryadok.ru до скидки_ИЦ2'] > df['МП Poryadok.ru до скидкиРРЦ']), 'МП Poryadok.ru до скидки_ИЦ_Последняя цена за период'] = df['МП Poryadok.ru до скидки_ИЦ2']
    # df.to_excel("12345.xlsx", index=False)

    
    # Рассчитываем "Оптовая 1" только для строк, где "изменение" является числом
    df['Оптовая 1'] = np.where(
        (~df['Оптовая 1_Последняя цена за период'].isna()) & (~df['Закупочная2'].isna()),  # Проверяем, что "изменение" и "Закупочная2" не NaN
        (df['Оптовая 1_Последняя цена за период'] / df['Закупочная2'] * 100) - 100,   # Выполняем расчёт
        np.nan)

    df['Розничная'] = np.where(
        (~df['Розничная_ИЦ_Последняя цена за период'].isna()) & (~df['Закупочная2'].isna()),  # Проверяем, что "изменение" и "Закупочная2" не NaN
        (df['Розничная_ИЦ_Последняя цена за период'] / df['Закупочная2'] * 100) - 100,   # Выполняем расчёт
        np.nan)

    df['МП Poryadok.ru до скидки'] = np.where(
        (~df['МП Poryadok.ru до скидки_ИЦ_Последняя цена за период'].isna()) & (~df['Закупочная2'].isna()),  # Проверяем, что "изменение" и "Закупочная2" не NaN
        (df['МП Poryadok.ru до скидки_ИЦ_Последняя цена за период'] / df['Закупочная2'] * 100) - 100,   # Выполняем расчёт
        np.nan)

    
    df['Розничная'] = (df['Розничная']).apply(custom_round_decimal)
    df['Оптовая 1'] = (df['Оптовая 1']).apply(custom_round_decimal)
    df['МП Poryadok.ru до скидки'] = (df['МП Poryadok.ru до скидки']).apply(custom_round_decimal)

    # Добавляем столбец с отметкой о наличии РРЦ
    df['Есть РРЦ'] = df['Оптовая_1РРЦ'].notna()
    


    df1 = df[~df['Оптовая 1'].isna()][['Код', 'Номенклатура', 'Оптовая 1', 'Есть РРЦ']]
    df2 = df[~df['Розничная'].isna()][['Код', 'Номенклатура', 'Розничная']]
    df3 = df[~df['МП Poryadok.ru до скидки'].isna()][['Код', 'Номенклатура', 'МП Poryadok.ru до скидки']]


    now1 = datetime.now()
    now1 = now1.strftime("%d-%m-%Y")

    # Функция для создания файла с подсветкой
    def create_highlighted_file(dataframe, file_name):
        # Создаем новую книгу
        wb = Workbook()
        ws = wb.active

        # Записываем данные из DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Применяем красную заливку для строк с РРЦ
        red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')

        # Находим индекс столбца 'Есть РРЦ' (последний столбец)
        rrc_column_idx = dataframe.shape[1]  # Индекс последнего столбца

        for row_idx in range(2, ws.max_row + 1):  # Пропускаем заголовок
            has_rrc = ws.cell(row=row_idx, column=rrc_column_idx).value
            if has_rrc:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = red_fill

        # Удаляем столбец 'Есть РРЦ'
        ws.delete_cols(rrc_column_idx)

        # Сохраняем файл
        wb.save(file_path + file_name + ' ' + now1 + '.xlsx')

    # Создаем файлы с подсветкой
    create_highlighted_file(df1, 'Для загрузки Оптовая 1')
    df2.to_excel(file_path + 'Для загрузки Розничная' + ' ' + now1 +'.xlsx', index=False)
    df3.to_excel(file_path + 'Для загрузки МП Poryadok.ru до скидки' + ' ' + now1 +'.xlsx', index=False)

    destination_folder = file_path + r'архив_отчетов'
    if not os.path.exists(destination_folder):
        os.makedirs(destination_folder)

    
    date_folder = os.path.join(destination_folder, now1)
    # Создание папки с текущей датой
    if not os.path.exists(date_folder):
        os.makedirs(date_folder)
    shutil.copy2(file_name_otchet, os.path.join(date_folder))
    

    add_message(f'Готово за {time() - t}')
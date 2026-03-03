
import pandas as pd
import os
import openpyxl
from time import time
from datetime import datetime
import sys
import numpy as np
from openpyxl.styles import Font
import glob
from python_calamine.pandas import pandas_monkeypatch


pandas_monkeypatch()

pd.options.mode.chained_assignment = None

os.chdir(os.path.dirname(os.path.abspath(__file__)))


def nalichie_po_formatam(file_path, add_message):
    if not file_path.endswith(os.sep):
        file_path = file_path + os.sep

    t = time()
    add_message('Стартуем!')
    
    d = {f: os.stat(f).st_mtime for f in glob.iglob(file_path +  "Исходники\\*.xlsx")}

    files = []
    for i in d.keys():
        if "сегмент" in i:
            continue
        else:
            files.append(i)

    df_1 = pd.read_excel(files[0], engine='calamine')
    df_2 = pd.read_excel(files[1], engine='calamine')

    # d1 = {f: os.stat(f).st_mtime for f in glob.iglob("Исходники\\*.csv")}
    # file_name_csv = max(d1, key=lambda i: d1[i])
    segment_file = file_path +  "Исходники\\сегмент.xlsx"

    df_main3 = pd.read_excel(segment_file,
                            usecols=['Код (доп.)', 'Подгруппа 1', 'СегментСтелажногоХранения'], engine='calamine')
    df_main3.rename(columns={'Код (доп.)': 'Код'}, inplace=True)
    df_main3['Код'] = pd.to_numeric(df_main3['Код'], errors='coerce')

    if 'Текущая наполненность ассортимента' in df_1.values:
        df_main1 = df_1.copy()
        df_main2 = df_2.copy()
    elif 'Остатки и доступность товаров' in df_1.values:
        df_main2 = df_1.copy()
        df_main1 = df_2.copy()

    
    add_message('Прочитали файлы')

    # обрабатываем файл текущей наполненности ассортимента
    i, j = np.where(df_main1.values == 'Код')
    df_main1 = df_main1.truncate(before=(i[0]))

    i0, s = np.where(df_main1.values == 'Ф1')
    i1, q = np.where(df_main1.values == 'Ф2')
    i2, r = np.where(df_main1.values == 'Ф3')
    i3, t = np.where(df_main1.values == 'Ф4')
    i4, y = np.where(df_main1.values == 'ФП')
    i5, x = np.where(df_main1.values == 'Собственная торговая марка')

    df_main1 = df_main1[[df_main1.columns[j[0]], df_main1.columns[s[0]], df_main1.columns[q[0]],
                            df_main1.columns[r[0]], df_main1.columns[t[0]], df_main1.columns[y[0]],
                            df_main1.columns[x[0]]]]
    df_main1.rename(columns=df_main1.iloc[0], inplace=True)
    df_main1.dropna(subset=['Код'], inplace=True)
    df_main1 = df_main1.iloc[1:]
    df_main1['Код'] = pd.to_numeric(df_main1['Код'], errors='coerce')

    # обрабатываем файл остатки
    i, j = np.where(df_main2.values == 'Сейчас')
    df_main2 = df_main2.truncate(before=(i[0]))

    i0, s = np.where(df_main2.values == 'Номенклатура.Код')
    df_main2 = df_main2.truncate(before=(i[0]))
    i1, q = np.where(df_main2.values == 'Доступно')
    df_main2 = df_main2[[df_main2.columns[s[0]], df_main2.columns[q[0]]]]
    df_main2.rename(columns=df_main2.iloc[1], inplace=True)
    df_main2 = df_main2.iloc[3:]
    df_main2.rename(columns={'Номенклатура.Код': 'Код'}, inplace=True)
    df_main2['Код'] = pd.to_numeric(df_main2['Код'], errors='coerce')

    df1 = df_main1.merge(df_main3, on='Код', how='left')
    df2 = df1.merge(df_main2, on='Код', how='left')
    df2['Доступно'].fillna(0, inplace=True)
    df2['СегментСтелажногоХранения'].fillna(99999, inplace=True)
    df2 = df2[df2['СегментСтелажногоХранения'] != 99999]

    add_message('Считаем наличие по сегментам >=10 шт')

    wb = openpyxl.Workbook()

    # наличие по сегментам
    # >=10
    sheet = wb.create_sheet("Наличие по сегментам >=10", 1)

    group = []
    for form1 in range(len(df2.index)):
        if df2['СегментСтелажногоХранения'].iloc[form1] in group:
            continue
        else:
            group.append(df2['СегментСтелажногоХранения'].iloc[form1])

    group = sorted(group)

    for i in range(0, len(group)):
        sheet['A' + str(i + 2)] = group[i]

    df_form = (df2[df2['Ф1'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 10])
    f1000_1 = (len(df_1.index) / len(df_form.index) * 100)

    c = 2
    for group1 in group:
        df_group = (df_form[df_form['СегментСтелажногоХранения'] == group1])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 10])
            sheet['B' + str(c)] = str(float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    df_form = (df2[df2['Ф2'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 10])
    f1300_1 = (len(df_1.index) / len(df_form.index) * 100)

    c = 2
    for group1 in group:
        df_group = (df_form[df_form['СегментСтелажногоХранения'] == group1])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 10])
            sheet['C' + str(c)] = str(float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    df_form = (df2[df2['Ф3'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 10])
    f1800_1 = (len(df_1.index) / len(df_form.index) * 100)

    c = 2
    for group1 in group:
        df_group = (df_form[df_form['СегментСтелажногоХранения'] == group1])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 10])
            sheet['D' + str(c)] = str(float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    df_form = (df2[df2['Ф4'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 10])
    f1800_1_1 = (len(df_1.index) / len(df_form.index) * 100)

    c = 2
    for group1 in group:
        df_group = (df_form[df_form['СегментСтелажногоХранения'] == group1])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 10])
            sheet['E' + str(c)] = str(float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    df_form = (df2[df2['ФП'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 10])
    f2000_1 = (len(df_1.index) / len(df_form.index) * 100)

    c = 2
    for group1 in group:
        df_group = (df_form[df_form['СегментСтелажногоХранения'] == group1])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 10])
            sheet['F' + str(c)] = str(float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    sheet["A1"] = 'Сегмент'
    sheet["B1"] = 'Ф1'
    sheet["C1"] = 'Ф2'
    sheet["D1"] = 'Ф3'
    sheet["E1"] = 'Ф4'
    sheet["F1"] = 'ФП'

    sheet.row_dimensions[1].height = 25
    sheet.column_dimensions['A'].width = 52

    # >=5
    add_message('Считаем наличие по сегментам >=5 шт')
    sheet = wb.create_sheet("Наличие по сегментам >=5", 2)

    group = []
    for form1 in range(len(df2.index)):
        if df2['СегментСтелажногоХранения'].iloc[form1] in group:
            continue
        else:
            group.append(df2['СегментСтелажногоХранения'].iloc[form1])

    group = sorted(group)

    for i in range(0, len(group)):
        sheet['A' + str(i + 2)] = group[i]

    df_form = (df2[df2['Ф1'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 5])
    f1000_1 = (len(df_1.index) / len(df_form.index) * 100)

    c = 2
    for group1 in group:
        df_group = (df_form[df_form['СегментСтелажногоХранения'] == group1])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 5])
            sheet['B' + str(c)] = str(float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    df_form = (df2[df2['Ф2'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 5])
    f1300_1 = (len(df_1.index) / len(df_form.index) * 100)

    c = 2
    for group1 in group:
        df_group = (df_form[df_form['СегментСтелажногоХранения'] == group1])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 5])
            sheet['C' + str(c)] = str(float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    df_form = (df2[df2['Ф3'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 5])
    f1800_1 = (len(df_1.index) / len(df_form.index) * 100)

    c = 2
    for group1 in group:
        df_group = (df_form[df_form['СегментСтелажногоХранения'] == group1])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 5])
            sheet['D' + str(c)] = str(float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    df_form = (df2[df2['Ф4'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 5])
    f1800_1_1 = (len(df_1.index) / len(df_form.index) * 100)

    c = 2
    for group1 in group:
        df_group = (df_form[df_form['СегментСтелажногоХранения'] == group1])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 5])
            sheet['E' + str(c)] = str(float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    df_form = (df2[df2['ФП'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 5])
    f2000_1 = (len(df_1.index) / len(df_form.index) * 100)

    c = 2
    for group1 in group:
        df_group = (df_form[df_form['СегментСтелажногоХранения'] == group1])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 5])
            sheet['F' + str(c)] = str(float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    sheet["A1"] = 'Сегмент'
    sheet["B1"] = 'Ф1'
    sheet["C1"] = 'Ф2'
    sheet["D1"] = 'Ф3'
    sheet["E1"] = 'Ф4'
    sheet["F1"] = 'ФП'

    sheet.row_dimensions[1].height = 25
    sheet.column_dimensions['A'].width = 52

    # >=1
    add_message('Считаем наличие по сегментам >=1 шт')
    sheet = wb.create_sheet("Наличие по сегментам >=1", 3)

    group = []
    for form1 in range(len(df2.index)):
        if df2['СегментСтелажногоХранения'].iloc[form1] in group:
            continue
        else:
            group.append(df2['СегментСтелажногоХранения'].iloc[form1])

    group = sorted(group)

    for i in range(0, len(group)):
        sheet['A' + str(i + 2)] = group[i]

    df_form = (df2[df2['Ф1'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 1])
    f1000_1 = (len(df_1.index) / len(df_form.index) * 100)

    c = 2
    for group1 in group:
        df_group = (df_form[df_form['СегментСтелажногоХранения'] == group1])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 1])
            sheet['B' + str(c)] = str(float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    df_form = (df2[df2['Ф2'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 1])
    f1300_1 = (len(df_1.index) / len(df_form.index) * 100)

    c = 2
    for group1 in group:
        df_group = (df_form[df_form['СегментСтелажногоХранения'] == group1])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 1])
            sheet['C' + str(c)] = str(float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    df_form = (df2[df2['Ф3'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 1])
    f1800_1 = (len(df_1.index) / len(df_form.index) * 100)

    c = 2
    for group1 in group:
        df_group = (df_form[df_form['СегментСтелажногоХранения'] == group1])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 1])
            sheet['D' + str(c)] = str(float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    df_form = (df2[df2['Ф4'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 1])
    f1800_1_1 = (len(df_1.index) / len(df_form.index) * 100)

    c = 2
    for group1 in group:
        df_group = (df_form[df_form['СегментСтелажногоХранения'] == group1])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 1])
            sheet['E' + str(c)] = str(float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    df_form = (df2[df2['ФП'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 1])
    f2000_1 = (len(df_1.index) / len(df_form.index) * 100)

    c = 2
    for group1 in group:
        df_group = (df_form[df_form['СегментСтелажногоХранения'] == group1])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 1])
            sheet['F' + str(c)] = str(float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    sheet["A1"] = 'Сегмент'
    sheet["B1"] = 'Ф1'
    sheet["C1"] = 'Ф2'
    sheet["D1"] = 'Ф3'
    sheet["E1"] = 'Ф4'
    sheet["F1"] = 'ФП'

    sheet.row_dimensions[1].height = 25
    sheet.column_dimensions['A'].width = 52

    # наличие по группам
    add_message('Считаем наличие по группам')

    sheet1 = wb.create_sheet("Наличие по группам", 0)

    group1 = []
    for form1 in range(len(df2.index)):
        if df2['Подгруппа 1'].iloc[form1] in group1:
            continue
        else:
            group1.append(df2['Подгруппа 1'].iloc[form1])

    group1 = sorted(group1)

    for i in range(0, len(group1)):
        sheet1['A' + str(i + 2)] = group1[i]

    df_form = (df2[df2['Ф1'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 10])
    f1_ost_10 = (len(df_1.index) / len(df_form.index) * 100)
    df_ost_5 = (df_form[df_form['Доступно'] >= 5])
    f1_ost_5 = (len(df_ost_5.index) / len(df_form.index) * 100)
    df_ost_1 = (df_form[df_form['Доступно'] >= 1])
    f1_ost_1 = (len(df_ost_1.index) / len(df_form.index) * 100)
    df_form_stm = df2[(df2['Ф1'] == 'Постоянный') & (df2['Собственная торговая марка'] == 'Да')]
    df_ost_stm = (df_form_stm[df_form_stm['Доступно'] >= 10])
    if len(df_form_stm.index) > 0:
        f1_ost_stm = (len(df_ost_stm.index) / len(df_form_stm.index)) * 100
    else:
        f1_ost_stm = 0

    c = 2
    for group2 in group1:
        df_group = (df_form[df_form['Подгруппа 1'] == group2])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 10])
            sheet1['B' + str(c)] = str(
                float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    sheet1['A' + str(c + 1)] = 'Итого с наличичем >= 10шт'
    sheet1['A' + str(c + 2)] = 'Итого с наличичем >= 5шт'
    sheet1['A' + str(c + 3)] = 'Итого с наличичем >= 1шт'
    sheet1['A' + str(c + 4)] = 'СТМ с наличичем >= 10шт'
    sheet1['A' + str(c + 1)].font = Font(bold=True)
    sheet1['A' + str(c + 2)].font = Font(bold=True)
    sheet1['A' + str(c + 3)].font = Font(bold=True)
    sheet1['A' + str(c + 4)].font = Font(bold=True)
    sheet1['B' + str(c + 1)] = str(float('{:.2f}'.format(f1_ost_10))) + ' %'
    sheet1['B' + str(c + 2)] = str(float('{:.2f}'.format(f1_ost_5))) + ' %'
    sheet1['B' + str(c + 3)] = str(float('{:.2f}'.format(f1_ost_1))) + ' %'
    sheet1['B' + str(c + 4)] = str(float('{:.2f}'.format(f1_ost_stm))) + ' %'

    df_form = (df2[df2['Ф2'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 10])
    f2_ost_10 = (len(df_1.index) / len(df_form.index) * 100)
    df_ost_5 = (df_form[df_form['Доступно'] >= 5])
    f2_ost_5 = (len(df_ost_5.index) / len(df_form.index) * 100)
    df_ost_1 = (df_form[df_form['Доступно'] >= 1])
    f2_ost_1 = (len(df_ost_1.index) / len(df_form.index) * 100)
    df_form_stm = df2[(df2['Ф2'] == 'Постоянный') & (df2['Собственная торговая марка'] == 'Да')]
    df_ost_stm = (df_form_stm[df_form_stm['Доступно'] >= 10])
    if len(df_form_stm.index) > 0:
        f2_ost_stm = (len(df_ost_stm.index) / len(df_form_stm.index)) * 100
    else:
        f2_ost_stm = 0

    c = 2
    for group2 in group1:
        df_group = (df_form[df_form['Подгруппа 1'] == group2])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 10])
            sheet1['C' + str(c)] = str(
                float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    sheet1['C' + str(c + 1)] = str(float('{:.2f}'.format(f2_ost_10))) + ' %'
    sheet1['C' + str(c + 2)] = str(float('{:.2f}'.format(f2_ost_5))) + ' %'
    sheet1['C' + str(c + 3)] = str(float('{:.2f}'.format(f2_ost_1))) + ' %'
    sheet1['C' + str(c + 4)] = str(float('{:.2f}'.format(f2_ost_stm))) + ' %'

    df_form = (df2[df2['Ф3'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 10])
    f3_ost_10 = (len(df_1.index) / len(df_form.index) * 100)
    df_ost_5 = (df_form[df_form['Доступно'] >= 5])
    f3_ost_5 = (len(df_ost_5.index) / len(df_form.index) * 100)
    df_ost_1 = (df_form[df_form['Доступно'] >= 1])
    f3_ost_1 = (len(df_ost_1.index) / len(df_form.index) * 100)
    df_form_stm = df2[(df2['Ф3'] == 'Постоянный') & (df2['Собственная торговая марка'] == 'Да')]
    df_ost_stm = (df_form_stm[df_form_stm['Доступно'] >= 10])
    if len(df_form_stm.index) > 0:
        f3_ost_stm = (len(df_ost_stm.index) / len(df_form_stm.index)) * 100
    else:
        f3_ost_stm = 0

    c = 2
    for group2 in group1:
        df_group = (df_form[df_form['Подгруппа 1'] == group2])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 10])
            sheet1['D' + str(c)] = str(
                float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    sheet1['D' + str(c + 1)] = str(float('{:.2f}'.format(f3_ost_10))) + ' %'
    sheet1['D' + str(c + 2)] = str(float('{:.2f}'.format(f3_ost_5))) + ' %'
    sheet1['D' + str(c + 3)] = str(float('{:.2f}'.format(f3_ost_1))) + ' %'
    sheet1['D' + str(c + 4)] = str(float('{:.2f}'.format(f3_ost_stm))) + ' %'

    df_form = (df2[df2['Ф4'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 10])
    f4_ost_10 = (len(df_1.index) / len(df_form.index) * 100)
    df_ost_5 = (df_form[df_form['Доступно'] >= 5])
    f4_ost_5 = (len(df_ost_5.index) / len(df_form.index) * 100)
    df_ost_1 = (df_form[df_form['Доступно'] >= 1])
    f4_ost_1 = (len(df_ost_1.index) / len(df_form.index) * 100)
    df_form_stm = df2[(df2['Ф4'] == 'Постоянный') & (df2['Собственная торговая марка'] == 'Да')]
    df_ost_stm = (df_form_stm[df_form_stm['Доступно'] >= 10])
    if len(df_form_stm.index) > 0:
        f4_ost_stm = (len(df_ost_stm.index) / len(df_form_stm.index)) * 100
    else:
        f4_ost_stm = 0

    c = 2
    for group2 in group1:
        df_group = (df_form[df_form['Подгруппа 1'] == group2])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 10])
            sheet1['E' + str(c)] = str(
                float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    sheet1['E' + str(c + 1)] = str(float('{:.2f}'.format(f4_ost_10))) + ' %'
    sheet1['E' + str(c + 2)] = str(float('{:.2f}'.format(f4_ost_5))) + ' %'
    sheet1['E' + str(c + 3)] = str(float('{:.2f}'.format(f4_ost_1))) + ' %'
    sheet1['E' + str(c + 4)] = str(float('{:.2f}'.format(f4_ost_stm))) + ' %'

    df_form = (df2[df2['ФП'] == 'Постоянный'])
    df_1 = (df_form[df_form['Доступно'] >= 10])
    f5_ost_10 = (len(df_1.index) / len(df_form.index) * 100)
    df_ost_5 = (df_form[df_form['Доступно'] >= 5])
    f5_ost_5 = (len(df_ost_5.index) / len(df_form.index) * 100)
    df_ost_1 = (df_form[df_form['Доступно'] >= 1])
    f5_ost_1 = (len(df_ost_1.index) / len(df_form.index) * 100)
    df_form_stm = df2[(df2['ФП'] == 'Постоянный') & (df2['Собственная торговая марка'] == 'Да')]
    df_ost_stm = (df_form_stm[df_form_stm['Доступно'] >= 10])
    if len(df_form_stm.index) > 0:
        f5_ost_stm = (len(df_ost_stm.index) / len(df_form_stm.index)) * 100
    else:
        f5_ost_stm = 0

    c = 2
    for group2 in group1:
        df_group = (df_form[df_form['Подгруппа 1'] == group2])
        if len(df_group.index) == 0:
            c += 1
            continue
        else:
            df_group1 = (df_group[df_group['Доступно'] >= 10])
            sheet1['F' + str(c)] = str(
                float('{:.2f}'.format((len(df_group1.index) / len(df_group.index) * 100)))) + ' %'
            c += 1

    sheet1['F' + str(c + 1)] = str(float('{:.2f}'.format(f5_ost_10))) + ' %'
    sheet1['F' + str(c + 2)] = str(float('{:.2f}'.format(f5_ost_5))) + ' %'
    sheet1['F' + str(c + 3)] = str(float('{:.2f}'.format(f5_ost_1))) + ' %'
    sheet1['F' + str(c + 4)] = str(float('{:.2f}'.format(f5_ost_stm))) + ' %'

    sheet1["A1"] = 'Старшая группа с наличием >= 10шт'
    sheet1["B1"] = 'Ф1'
    sheet1["C1"] = 'Ф2'
    sheet1["D1"] = 'Ф3'
    sheet1["E1"] = 'Ф4'
    sheet1["F1"] = 'ФП'
    sheet1["A1"].font = Font(bold=True)
    sheet1["B1"].font = Font(bold=True)
    sheet1["C1"].font = Font(bold=True)
    sheet1["D1"].font = Font(bold=True)
    sheet1["E1"].font = Font(bold=True)
    sheet1["F1"].font = Font(bold=True)

    sheet1.row_dimensions[1].height = 25
    sheet1.column_dimensions['A'].width = 52

    # Список имен вкладок, которые нужно сохранить
    worksheet_names = ['Наличие по группам', 'Наличие по сегментам >=10', 'Наличие по сегментам >=5',
                        'Наличие по сегментам >=1']

    # Удаление вкладок, которые не нужно сохранять
    for worksheet in wb.worksheets:
        if worksheet.title not in worksheet_names:
            wb.remove(worksheet)

    now1 = datetime.now()
    now1 = now1.strftime("%d-%m-%Y")

    wb.save(file_path + "Наличие итог " + now1 + ".xlsx")

    add_message(f'Готово за {time() - t}')




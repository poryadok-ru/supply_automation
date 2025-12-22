import sys
import pandas as pd
import numpy as np
import glob
import os
import shutil
from time import time
from datetime import datetime, timedelta
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import xml.etree.ElementTree as ET
from pandas import read_excel
from python_calamine.pandas import pandas_monkeypatch


pandas_monkeypatch()



pd.options.mode.chained_assignment = None


os.chdir(os.path.dirname(os.path.abspath(__file__)))


def optzakazfivew(file_path, checkbox_CopyToOrders, checkbox_nul_3tr, checkbox_not_nul_5tr_menee_pol_up, porog_max, porog_min, add_message):
    if not file_path.endswith(os.sep):
        file_path = file_path + os.sep

    t = time()
    add_message('Стартуем!')
    add_message('Определяем период продаж')

    if porog_max < 1000:
        porog_max = 1000
    if porog_min < 900:
        porog_min = 900
    

    # определяем названия колонок с продажами и добавляем их в список months
    
    now = datetime.now() - timedelta(days=1)

    months = ['Продажи за ' + now.strftime("%m-%Y")]
    month, year = now.month, now.year
    for i in range(0, 12):
        month1, year1 = (month - 1, year) if month != 1 else (12, year - 1)
        months.append('Продажи за ' + now.replace(day=1, month=month1, year=year1).strftime("%m-%Y"))
        month, year = month1, year1
    
    add_message('Считываем файлы')

    
    d = {f: os.stat(f).st_mtime for f in glob.iglob(file_path + "файлы_заказов\\*.xlsx")}
    files = []
    for i in d.keys():
        files.append(i)

    itog_folder = file_path + r'итоговые_файлы'
    for filename in os.listdir(itog_folder):
        file_path_del = os.path.join(itog_folder, filename)
        if os.path.isfile(file_path_del):
            os.remove(file_path_del)

    for k in range(0, len(files)):
        df_af = pd.read_excel(files[k], engine="calamine")
        if 'Волгоградская АВС (ШафиевТ)' in df_af.columns:
            df_af.rename(columns={'Волгоградская АВС (ШафиевТ)': 'Волгоградская АВС'}, inplace=True)
        if 'Подольск АВС (ШафиевТ)' in df_af.columns:
            df_af.rename(columns={'Подольск АВС (ШафиевТ)': 'Подольск АВС'}, inplace=True)
        df_af['Комментарий'] = ''

        n = {f: os.stat(f).st_mtime for f in glob.iglob(r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Шафиев\Остатки_розница\\*.xlsx")}
        file_name_ost = max(n, key=lambda i: n[i])
        ost_rozn_column = 'Ост Розн+FBO с уч аналогов ' + datetime.fromtimestamp(n[file_name_ost]).strftime('%d-%m-%Y')
        # print(ost_rozn_column)
        df_OST_ROZN = pd.read_excel(file_name_ost, engine="calamine")
        df_OST_ROZN = df_OST_ROZN.rename(columns={'Розница с учетом аналогов': ost_rozn_column})
        # print(df_OST_ROZN.columns)
        df_af = df_af.merge(df_OST_ROZN, on='Артикул (доп)', how='left')
        df_af[ost_rozn_column] = df_af[ost_rozn_column].fillna(0)

        df_af = df_af[['Артикул (доп)', 'Продукт', ost_rozn_column, 'В резерве',
                        'Фактический остаток', 'Заказано', 'Оптимальный запас', 'Заказать',
                        'Цена закупа', 'Сумма', 'УП1, ед.', 'УП2, ед.',
                        'Горизонт планирования', 'Срок доставки в днях',
                        'Волгоградская АВС', 'Склад(Название)'] +
                        months + ['Округление', 'Округление (УП2)', 'ЗакупочнаяЦена', 'Заказать без обработок', 'Комментарий',
                                'Поставщик для заказа (Название)', 'Артикул', 'Период действия будущей акции', 'Период действия текущей акции', 'Вес']]
        df_af = df_af.dropna(subset=['Продукт'])

        df_af['Артикул (доп)'] = pd.to_numeric(df_af['Артикул (доп)'], errors='coerce')
        df_af['УП1, ед.'] = pd.to_numeric(df_af['УП1, ед.'], errors='coerce')
        df_af['ЗакупочнаяЦена'] = pd.to_numeric(df_af['ЗакупочнаяЦена'], errors='coerce')
        df_af['ЗакупочнаяЦена'].fillna(0, inplace=True)

        df_af['ОСТАТОК'] = df_af['Фактический остаток'] + df_af['Заказано'] - df_af['В резерве']

        # если цена закупа = 0, то приравниваем к закупочной (если есть)
        df_af.loc[((df_af['Цена закупа'] == 0) &
                    (df_af['ЗакупочнаяЦена'] != 0)), 'Цена закупа'] = df_af['ЗакупочнаяЦена']
        

        # ### Округление

        # # определение приоритетного окргуления УП2 или УП1

        # okruglenie_up2 = df_af['Округление (УП2)'].unique()[0]

        # if okruglenie_up2 != 'Не установлено':
        #     # если заказать без обработок >= 0.5*УП2

        #     df_af.loc[(df_af['Округление (УП2)'] == 'По правилам') &
        #                 (df_af['Заказать'] % df_af['УП2, ед.'] >= 0.5 * df_af['УП2, ед.']),
        #                 'Заказать'] = (df_af['Заказать'] // df_af['УП2, ед.'] + 1) * df_af['УП2, ед.']
            
        #     # если заказать без обработок < 0.5*УП2 с условием по остатку
            
        #     df_af.loc[(df_af['Округление (УП2)'] == 'По правилам') &
        #                 (df_af['Заказать'] % df_af['УП2, ед.'] < 0.5 * df_af['УП2, ед.']) &
        #                 (df_af['Заказать'] > 0),
        #                 'Заказать'] = (df_af['Заказать'] // df_af['УП2, ед.']) * df_af['УП2, ед.']
            
        #     ### Правило - половина кратности, складская

        #     df_af.loc[(df_af['Округление (УП2)'] == 'Половина кратности, складская') &
        #                 (df_af['Заказать'] % df_af['УП2, ед.'] >= 0.5 * df_af['УП2, ед.']),
        #                 'Заказать'] = (df_af['Заказать'] // df_af['УП2, ед.'] + 1) * df_af['УП2, ед.']
            
        #     # для заказа больше одной кратности
            
        #     df_af.loc[(df_af['Округление (УП2)'] == 'Половина кратности, складская') &
        #                 (df_af['Заказать'] % df_af['УП2, ед.'] < 0.5 * df_af['УП2, ед.']) &
        #                 ((df_af['Заказать'] // df_af['УП2, ед.']) >= 1),
        #                 'Заказать'] = (df_af['Заказать'] // df_af['УП2, ед.']) * df_af['УП2, ед.']
            
        #     # для заказа меньше одной кратности
            
        #     df_af.loc[(df_af['Округление (УП2)'] == 'Половина кратности, складская') &
        #                 (df_af['Заказать'] % df_af['УП2, ед.'] < 0.5 * df_af['УП2, ед.']) &
        #                 ((df_af['Заказать'] // df_af['УП2, ед.']) == 0),
        #                 'Заказать'] = df_af['Заказать']

        # else:
        #     df_af.loc[(df_af['Округление'] == 'По правилам') &
        #                 (df_af['Заказать'] % df_af['УП1, ед.'] >= 0.5 * df_af['УП1, ед.']),
        #                 'Заказать'] = (df_af['Заказать'] // df_af['УП1, ед.'] + 1) * df_af['УП1, ед.']
        #     # print(df_af['Заказать'])
            
        #     # если заказать без обработок < 0.5*УП2 с условием по остатку
            
        #     df_af.loc[(df_af['Округление'] == 'По правилам') &
        #                 (df_af['Заказать'] % df_af['УП1, ед.'] < 0.5 * df_af['УП1, ед.']) &
        #                 (df_af['Заказать'] > 0),
        #                 'Заказать'] = (df_af['Заказать'] // df_af['УП1, ед.']) * df_af['УП1, ед.']
        #     # print(df_af['Заказать'])
            
        #     ### Правило - половина кратности, складская

        #     df_af.loc[(df_af['Округление'] == 'Половина кратности, складская') &
        #                 (df_af['Заказать'] % df_af['УП1, ед.'] >= 0.5 * df_af['УП1, ед.']),
        #                 'Заказать'] = (df_af['Заказать'] // df_af['УП1, ед.'] + 1) * df_af['УП1, ед.']
            
        #     # для заказа больше одной кратности
            
        #     df_af.loc[(df_af['Округление'] == 'Половина кратности, складская') &
        #                 (df_af['Заказать'] % df_af['УП1, ед.'] < 0.5 * df_af['УП1, ед.']) &
        #                 ((df_af['Заказать'] // df_af['УП1, ед.']) >= 1),
        #                 'Заказать'] = (df_af['Заказать'] // df_af['УП1, ед.']) * df_af['УП1, ед.']
            
        #     # для заказа меньше одной кратности
            
        #     df_af.loc[(df_af['Округление'] == 'Половина кратности, складская') &
        #                 (df_af['Заказать'] % df_af['УП1, ед.'] < 0.5 * df_af['УП1, ед.']) &
        #                 ((df_af['Заказать'] // df_af['УП1, ед.']) == 0),
        #                 'Заказать'] = df_af['Заказать']
            
        if 'Русбытхим ПК ООО_растворители' in files[k]:
            # print(np.ceil(df_af['Заказать'] / (10 * df_af['УП2, ед.'])) * (10 * df_af['УП2, ед.']))
            print(df_af['Заказать'] / (10 * df_af['УП2, ед.']))
            df_af.loc[((df_af['Заказать'] > 0) & (df_af['Заказать'] % (10 * df_af['УП2, ед.']) != 0)), 'Заказать'] = \
            np.ceil(df_af['Заказать'] / (10 * df_af['УП2, ед.'])) * (10 * df_af['УП2, ед.'])


        
        
        # увеличиваем до 1,5тр
        df_af.loc[((df_af['Заказать'] * df_af['Цена закупа']) <= porog_max) &
                    ((df_af['Заказать без обработок']) > 0) &
                    (df_af['Заказать'] == 0) &
                    (df_af['ОСТАТОК'] == 0) &
                    (df_af['Цена закупа'] != 0), 'Заказ с ↑ до ' + str(porog_max)] = \
            np.ceil(porog_max / df_af['Цена закупа'])
        
        
        # 
        df_af.loc[((df_af['Заказать'] * df_af['Цена закупа']) <= porog_max) &
                    ((df_af['Заказать без обработок']) > 0) &
                    ((df_af['Заказать'] == 0) |
                    (df_af['ОСТАТОК'] == 0)), 'Заказ с ↑ до ' + str(porog_max)] = 0

        df_af.loc[((df_af['Заказать'] * df_af['Цена закупа']) <= porog_max) &
                    ((df_af['Заказать']) > 0) &
                    (df_af['Цена закупа'] != 0), 'Заказ с ↑ до ' + str(porog_max)] = \
            np.ceil(porog_max / df_af['Цена закупа'])

        df_af.loc[((df_af['Заказать'] * df_af['Цена закупа']) <= porog_max) &
                    ((df_af['Заказать']) > 0), 'Комментарий'] = 'Увеличение до 1,5тр'


        # для правила по правилам
        df_af.loc[(((df_af['Заказать'] * df_af['Цена закупа']) <= porog_max) &
                    (df_af['Заказать'] > 0) &
                    (df_af['Заказ с ↑ до ' + str(porog_max)] >= (0.5 * df_af['УП1, ед.'])) &
                    (df_af['Округление'] == 'По правилам') &
                    (df_af['Цена закупа'] != 0)), 'Заказ с ↑ до ' + str(porog_max)] = \
            np.ceil(porog_max / df_af['Цена закупа'] / df_af['УП1, ед.']) * df_af['УП1, ед.']

        df_af.loc[(((df_af['Заказать'] * df_af['Цена закупа']) <= porog_max) &
                    (df_af['Заказать'] > 0) &
                    (df_af['Заказ с ↑ до ' + str(porog_max)] >= (0.5 * df_af['УП1, ед.'])) &
                    (df_af['Округление'] == 'По правилам') &
                    (df_af['Цена закупа'] != 0)), 'Комментарий'] = 'Увеличение до 1,5тр кратно уп'
        
        if checkbox_not_nul_5tr_menee_pol_up != 'on':
            # add_message('обнуляем заказ менее половины УП с суммой 5тр при округлении по правилам')
            df_af.loc[((df_af['Заказать'] * df_af['Цена закупа']) <= porog_max) &
                        (df_af['Заказать'] > 0) &
                        (df_af['Заказ с ↑ до ' + str(porog_max)] < (0.5 * df_af['УП1, ед.'])) &
                        (df_af['Округление'] == 'По правилам'), 'Заказ с ↑ до ' + str(porog_max)] = 0

            df_af.loc[((df_af['Заказать'] * df_af['Цена закупа']) <= porog_max) &
                        (df_af['Заказать'] > 0) &
                        (df_af['Заказ с ↑ до ' + str(porog_max)] < (0.5 * df_af['УП1, ед.'])) &
                        (df_af['Округление'] == 'По правилам'), 'Комментарий'] = 'Обнулен, заказ на 1,5тр менее половины уп'
        else:
            df_af.loc[((df_af['Заказать'] * df_af['Цена закупа']) <= porog_max) &
                        (df_af['Заказать'] > 0) &
                        (df_af['Заказ с ↑ до ' + str(porog_max)] < (0.5 * df_af['УП1, ед.'])) &
                        (df_af['Округление'] == 'По правилам'), 'Комментарий'] = 'ЗАКАЗ МОЖЕТ БЫТЬ НЕ КРАТЕН УП'

        # для правила половина кратности
        df_af.loc[((df_af['Заказать'] * df_af['Цена закупа']) <= porog_max) &
                    (df_af['Заказать'] > 0) &
                    (df_af['Заказ с ↑ до ' + str(porog_max)] >= (0.5 * df_af['УП1, ед.'])) &
                    (df_af['Округление'] == 'Половина кратности, складская') &
                    (df_af['Цена закупа'] != 0), 'Заказ с ↑ до ' + str(porog_max)] = \
            np.ceil(porog_max / df_af['Цена закупа'] / df_af['УП1, ед.']) * df_af['УП1, ед.']

        df_af.loc[((df_af['Заказать'] * df_af['Цена закупа']) <= porog_max) &
                    (df_af['Заказать'] > 0) &
                    (df_af['Заказ с ↑ до ' + str(porog_max)] >= (0.5 * df_af['УП1, ед.'])) &
                    (df_af['Округление'] == 'Половина кратности, складская') &
                    (df_af['Цена закупа'] != 0), 'Комментарий'] = \
            'Увеличение до 1,5тр кратно уп'
        
        if checkbox_nul_3tr == 'on':
            # add_message('Обнулили заказ менее porog_min р по строке')
            df_af.loc[((df_af['Заказать'] * df_af['Цена закупа']) < porog_min) &
                        (df_af['Цена закупа'] != 0) &
                        ((df_af['Заказать']) > 0) &
                        (df_af['ОСТАТОК'] > 1), 'Заказ с ↑ до ' + str(porog_max)] = 0

            df_af.loc[((df_af['Заказать'] * df_af['Цена закупа']) < porog_min) &
                        (df_af['Цена закупа'] != 0) &
                        ((df_af['Заказать']) > 0) &
                        (df_af['ОСТАТОК'] > 1), 'Комментарий'] = 'Обнуление заказа менее porog_minр'

        # df_af.loc[((df_af['Заказать'] * df_af['Цена закупа']) <= 5000) &
        #           (df_af['Заказать без обработок'] > 0) &
        #           (df_af['Заказ с ↑ до 5тр'] < (0, 5 * df_af['УП1, ед.'])) &
        #           (df_af['Округление'] == 'Половина кратности, складская'), 'Заказ с ↑ до 5тр'] = 0

        
        df_af.loc[((df_af['Заказать'] * df_af['Цена закупа']) > porog_max) |
                    (df_af['Цена закупа'] == 0), 'Заказ с ↑ до ' + str(porog_max)] = df_af['Заказать']
        df_af.loc[((df_af['Заказать без обработок'] == 0) &
                    (df_af['Заказать'] == 0)), 'Заказ с ↑ до ' + str(porog_max)] = df_af['Заказать']


        skl = df_af.loc[0, 'Склад(Название)']
        df_af['Вес суммарный'] = 0

        df_af = df_af[['Артикул (доп)', 'Продукт', ost_rozn_column, 'Фактический остаток', 'В резерве', 'Заказано',
                        'ОСТАТОК', 'Заказать', 'Заказ с ↑ до ' + str(porog_max), 'Цена закупа',
                        'Сумма', 'УП1, ед.', 'УП2, ед.', 'Горизонт планирования', 'Срок доставки в днях',
                        'Оптимальный запас', 'Волгоградская АВС', 'Склад(Название)'] + months +
                        ['Округление', 'Округление (УП2)', 'ЗакупочнаяЦена', 'Заказать без обработок', 'Комментарий',
                        'Поставщик для заказа (Название)', 'Артикул', 'Период действия будущей акции', 'Период действия текущей акции', 'Вес', 'Вес суммарный']]

        df_af.sort_values(by=['Продукт'], inplace=True)

        
        itog_path = os.path.join(itog_folder, os.path.basename(files[k]))

        df_af.to_excel(itog_path, index=False)
        wb = load_workbook(itog_path)
        ws = wb.active

        ws.insert_rows(1)

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['O'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['R'].width = 12
        ws.row_dimensions[2].height = 35

        ws.column_dimensions['L'].width = 5
        ws.column_dimensions['M'].width = 5
        ws.column_dimensions['N'].width = 9
        ws.column_dimensions['O'].width = 7

        # Перебираем все ячейки в первой строке
        for row in ws.iter_rows(min_row=2, max_row=2):
            for cell in row:
                # Проверка, содержит ли ячейка словосочетание "продажи за"
                if "Продажи за" in str(cell.value):
                    ws.column_dimensions[cell.column_letter].width = 5

        for row in ws.iter_rows(min_row=2, max_row=2):
            for cell in row:
                # Проверка, содержит ли ячейка словосочетание "продажи за"
                if "Продажи за" in cell.value:
                    # Разделение строки на две части по разделителю "-"
                    cell.value = cell.value.replace("Продажи за ", "")


        for cell in ws[2]:
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        c = str(len(df_af) + 2)
        ws['K1'] = '=SUM(K3:K' + c + ')'
        ws['K1'].number_format = '#,##0.00'
        ws['N1'] = '=SUM(AP3:AP' + c + ')'
        ws['N1'].number_format = '#,##0.00'
        ws['J1'] = 'Сумма:'
        ws['M1'] = 'Вес:'
        for i in range(3, (len(df_af) + 3)):
            ws['K' + str(i)] = '=J%s*I%s' % (i, i)
            ws['AP' + str(i)] = '=AO%s*I%s' % (i, i)
            # ws['J' + str(i)].number_format = '#,##0.00'
            ws['G' + str(i)].number_format = '#,##0'
            ws['H' + str(i)].number_format = '#,##0'
            ws['I' + str(i)].number_format = '#,##0'
            ws['AP' + str(i)].number_format = '#,##0'


        green_fill = PatternFill(start_color="bae5d1", end_color="bae5d1", fill_type="solid")
        rule1 = Rule(type="expression", dxf=DifferentialStyle(fill=green_fill))
        rule1.formula = ["AND(H3>0)"]
        ws.conditional_formatting.add(('H3:H' + c), rule1)
        rule11 = Rule(type="expression", dxf=DifferentialStyle(fill=green_fill))
        rule11.formula = ["AND(I3>0)"]
        ws.conditional_formatting.add(('I3:I' + c), rule11)

        pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
        rule = Rule(type="expression", dxf=DifferentialStyle(fill=pink_fill))
        if skl == 'Александровка':
            rule.formula = [f"AND(K3>=0.1, K3<{porog_max})"]
        elif skl == 'Подольск':
            rule.formula = ["AND(K3>=0.1, K3<{porog_max})"]

        ws.conditional_formatting.add(('K3:K' + c), rule)

        rule2 = Rule(type="expression", dxf=DifferentialStyle(fill=pink_fill))
        rule2.formula = ["G3<=2"]
        ws.conditional_formatting.add(('G3:G' + c), rule2)


        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

        font = Font(name='Arial', size=8)

        for row in ws.iter_rows():
            for cell in row:
                cell.font = font

        ws.freeze_panes = 'D3'
        itog_folder = file_path + r'итоговые_файлы'
        itog_path = os.path.join(itog_folder, os.path.basename(files[k]))

        wb.save(itog_path)

        add_message('Готов ' + os.path.basename(files[k]).rsplit(" ", 1)[0])

    excel_folder = itog_folder
    xml_folder = file_path + r'файлы_заказов'

    # Перебор файлов в папке "эксель"
    for excel_file in os.listdir(excel_folder):
        if excel_file.endswith('.xlsx'):
            excel_path = os.path.join(excel_folder, excel_file)
            xml_file = excel_file.replace('.xlsx', '.xml')
            xml_path = os.path.join(xml_folder, xml_file)

            # Чтение данных из Excel
            df = pd.read_excel(excel_path, header=1, engine="calamine")
            df['Заказ с ↑ до ' + str(porog_max)] = pd.to_numeric(df['Заказ с ↑ до ' + str(porog_max)], errors='coerce')
            # print(df)

            # Чтение и изменение соответствующего XML файла
            tree = ET.parse(xml_path)
            root = tree.getroot()

            for index, row in df.iterrows():
                code = str(row['Артикул (доп)'])
                quantity = str(row['Заказ с ↑ до ' + str(porog_max)])

                for elem in root.iter('row'):
                    if elem.find('rc-article-ext').text == code:
                        elem.find('rc-to-buy').text = quantity

            xml_path_itog = os.path.join(itog_folder, xml_file)

            # Сохранение изменений в XML файле с сохранением оригинальной кодировки
            tree.write(xml_path_itog, encoding='utf-8', xml_declaration=True)
        
    pattern_rms = glob.glob(os.path.join(itog_folder, '*Компания РМС ООО*.xml'))
    
    if len(pattern_rms) > 1:
        arm_file = next((file for file in pattern_rms if 'Компания РМС ООО_фитинги' in file), None)
        smes_file = next((file for file in pattern_rms if 'Компания РМС ООО_смесители' in file), None)
        output_file = os.path.join(itog_folder, 'Компания РМС ООО_смесители.xml')
        tree_smes = ET.parse(smes_file)
        root_smes = tree_smes.getroot()

        # Извлекаем значения <rc-supplier-id> и <rc-supplier-name>
        rc_supplier_id_smes = root_smes.find('.//rc-supplier-id').text if root_smes.find('.//rc-supplier-id') is not None else None
        rc_supplier_name_smes = root_smes.find('.//rc-supplier-name').text if root_smes.find('.//rc-supplier-name') is not None else None

        print(f"Извлеченные значения из файла фитинги: <rc-supplier-id> = {rc_supplier_id_smes}, <rc-supplier-name> = {rc_supplier_name_smes}")

        # Сохраняем блок <total> отдельно
        total_element = root_smes.find('total')

            # Удаляем блок <total> из первого файла
        if total_element is not None:
            root_smes.remove(total_element)
            
        # Находим контейнер для строк <row> в первом файле
        rows_container = root_smes
        

        # Находим максимальное значение <row-id> в первом файле
        max_row_id = 0
        for row in rows_container.findall('row'):
            row_id_element = row.find('row-id')
            if row_id_element is not None:
                row_id = int(row_id_element.text)
                if row_id > max_row_id:
                    max_row_id = row_id

        print(f"Максимальное значение <row-id> в первом файле: {max_row_id}")
        print("123")

        # Загружаем файл "Эльф ГК ООО_радиаторы"
        tree_arm = ET.parse(arm_file)
        root_arm = tree_arm.getroot()

        # Добавляем элементы <row> из второго файла
        for row in root_arm.findall('row'):
            new_row_id = max_row_id + 1
            row_id_element = row.find('row-id')

            if row_id_element is not None:
                row_id_element.text = str(new_row_id)
            else:
                row_id_element = ET.Element('row-id')
                row_id_element.text = str(new_row_id)
                row.append(row_id_element)

            # Заменяем <rc-supplier-id> и <rc-supplier-name> на значения из файла фитинги
            rc_supplier_id_element = row.find('rc-supplier-id')
            rc_supplier_name_element = row.find('rc-supplier-name')

            if rc_supplier_id_element is not None and rc_supplier_id_smes is not None:
                rc_supplier_id_element.text = rc_supplier_id_smes
            if rc_supplier_name_element is not None and rc_supplier_name_smes is not None:
                rc_supplier_name_element.text = rc_supplier_name_smes

            rows_container.append(row)
            max_row_id = new_row_id
        
        # Перемещаем блок <total> в конец
        if total_element is not None:
            root_smes.append(total_element)

        # Удаляем исходные файлы
        for file in pattern_rms:
            os.remove(file)
        print("Исходные файлы удалены.")
        

        # Сохраняем объединённый XML в новый файл
        tree_smes.write(output_file, encoding='utf-8', xml_declaration=True)

    target_folder = r'\\SRV-FNOW\Exch\Orders'

    if checkbox_CopyToOrders == 'on':
        add_message('Скопировали в Orders')
        for filename in os.listdir(itog_folder):
            if filename.endswith('.xml'):
                file_path_src = os.path.join(itog_folder, filename)
                file_path_dst = os.path.join(target_folder, filename)
                shutil.copy2(file_path_src, file_path_dst)
    
    source_folder = xml_folder
    destination_folder = file_path + r'архив_заказов'
    
    for filename in os.listdir(source_folder):
        if filename.endswith('.xlsx'):
            source_file_path = os.path.join(source_folder, filename)
            if os.path.isfile(source_file_path):
                destination_file_path = os.path.join(destination_folder, filename)
                file_path_dst2 = r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\В.Гречушкин\Общая\Доработки\График заказов"
                shutil.copy2(source_file_path, destination_file_path)
                shutil.copy2(source_file_path, file_path_dst2)

    for filename in os.listdir(excel_folder):
        if filename.endswith('.xlsx'):
            excel_folder_path = os.path.join(excel_folder, filename)
            if os.path.isfile(excel_folder_path):
                destination_file_path = os.path.join(destination_folder, filename)
                file_path_dst3 = r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\Заказы поставщикам"
                shutil.copy2(excel_folder_path, file_path_dst3)

    for filename in os.listdir(xml_folder):
        file_path_del = os.path.join(xml_folder, filename)
        if os.path.isfile(file_path_del):
            os.remove(file_path_del)


    add_message(f'Готово за {time() - t}')
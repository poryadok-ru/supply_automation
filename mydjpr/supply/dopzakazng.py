import pandas as pd
import numpy as np
import os
import glob
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill
from openpyxl.styles.numbers import NumberFormat
from openpyxl import load_workbook
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from pandas import read_excel
from python_calamine.pandas import pandas_monkeypatch


pandas_monkeypatch()


pd.options.mode.chained_assignment = None

os.chdir(os.path.dirname(os.path.abspath(__file__)))


def dop_ng(file_path, add_message):
    if not file_path.endswith(os.sep):
        file_path = file_path + os.sep
        
    df_sezon = pd.read_excel(r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Шафиев\Сезоны_ручной просмотр перед НГ.xlsx", engine="calamine")

    files = glob.glob(file_path + "Исходники\\*.csv")

    df_effectiv = pd.DataFrame()

    df_action = pd.read_csv(r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Шафиев\акции.csv", delimiter=';', low_memory=False, thousands=" ",
                            usecols=['Артикул(доп.)', 'Вид акции', 'Период акции'])

    df_action.sort_values(by=['Вид акции'])

    df_pri = pd.read_excel(r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Шафиев\кприрост.xlsx", engine="calamine")

    # Группировка и объединение значений
    df_action = df_action.groupby('Артикул(доп.)').agg({
        'Вид акции': lambda x: ', '.join(x.unique()),  # Объединяем уникальные виды акций
        'Период акции': lambda x: ', '.join(x.unique())  # Объединяем уникальные периоды акций
    }).reset_index()

    # Выводим результат
    # print(result)
    #
    # result.to_excel("123.xlsx", index=False)

    for k in range(0, len(files)):
        df_af = pd.read_csv(files[k], delimiter=';', low_memory=False, thousands=" ",
                            usecols=['Артикул(доп.)', 'Наименование', 'В', 'Остаток свободный (Из)', 'ЗакупочнаяЦена',
                                    'Количество для перемещения',
                                    'КомментарийТО', 'Остаток свободный (В)', 'Заказано (В)', 'ОТЗ (В) на норму запаса',
                                    'Кратность ,ед. (В)', 'ЗакупочнаяЦена', 'СегментСтелажногоХранения', 'Сезон', ''
                                    'Каталог (В)', 'Маячки (В)',
                                    'Период действия будущей акции (В)', 'Период действия текущей акции (В)',
                                    'Продажи за 12-2023 (В)', 'Суммарные продажи за 7 дней (В)', 'Объем'])

        df_effectiv = pd.concat([df_effectiv, df_af])


    # df_it = df_effectiv[df_effectiv['Сезон'].isin(df_sezon['Сезон'])]
    df_it = df_effectiv
    prk_number = df_it.iloc[1]['В']
    df_it['ЗАКАЗ'] = " "
    df_it['Сумма заказа'] = " "
    df_it['Объем заказа'] = " "
    df_it['Остаток+заказано'] = df_it['Остаток свободный (В)'] + df_it['Заказано (В)']
    df_it['Докинуть'] = " "
    df_it['ЗАКАЗ'] = pd.to_numeric(df_it['ЗАКАЗ'], errors='coerce')
    

    df_it = df_it.sort_values(by=['Сезон', 'Наименование'])
    df_it = df_it[df_it['Остаток свободный (Из)'] > 5]
    df_it = df_it[['Артикул(доп.)', 'Наименование', 'В', 'Остаток свободный (Из)', 'Суммарные продажи за 7 дней (В)',
                'Продажи за 12-2023 (В)', 'Остаток+заказано', 'Количество для перемещения', 'Докинуть', 'ЗАКАЗ',
                'ОТЗ (В) на норму запаса', 'Кратность ,ед. (В)',
                'Сумма заказа', 'Объем заказа', 'Каталог (В)', 'Маячки (В)', 'Период действия будущей акции (В)',
                'Период действия текущей акции (В)', 'Сезон', 'СегментСтелажногоХранения',
                    'ЗакупочнаяЦена', 'Объем']]

    df_it = df_it.merge(df_action, on='Артикул(доп.)', how='left')
    df_it = df_it.merge(df_pri, on=['В', 'СегментСтелажногоХранения'], how='left')

    now1 = datetime.now()
    now1 = now1.strftime("%d-%m-%Y")

    df_it.to_excel(file_path + 'Заказ ' + prk_number + ' ' + now1 + '.xlsx', index=False)

    wb = openpyxl.load_workbook(file_path + 'Заказ ' + prk_number + ' ' + now1 + '.xlsx')
    sheet = wb.active
    sheet.insert_rows(0, amount=1)
    sheet.row_dimensions[2].height = 70
    sheet.column_dimensions['B'].width = 43
    sheet.column_dimensions['C'].width = 13
    sheet.column_dimensions['Q'].width = 16
    sheet.column_dimensions['R'].width = 16
    sheet.column_dimensions['S'].width = 29
    sheet.column_dimensions['W'].width = 20
    sheet.column_dimensions['X'].width = 43
    for row in sheet.iter_rows(max_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    for i in range(2, (len(df_it)+3)):
        sheet['J'+str(i)].fill = PatternFill(fill_type='solid', start_color='00FFFF99')

    for i in range(3, (len(df_it)+3)):
        sheet['M'+str(i)] = '=J%s*U%s' % (i, i)
        sheet['N' + str(i)] = '=J%s*V%s' % (i, i)
        # sheet['I' + str(i)] = '=ROUND($C$1*F%s-G%s, 0)' % (i, i)
        # sheet['I' + str(i)] = (
        #     '=IF(ROUND(E{0}*Z{0}-G{0}, 0) < 0, '
        #     '0, ROUND(E{0}*Z{0}-G{0}, 0))'
        #     ).format(i)
        # sheet['I' + str(i)] = " "


    c = str(len(df_it)+2)
    sheet['M1'] = '=SUM(M3:M' + c +')'
    sheet['N1'] = '=SUM(N3:N' + c +')'
    sheet['B1'] = 'Коэффициент от продаж 12.23'
    sheet['C1'] = 1


    green_fill = PatternFill(start_color="bae5d1", end_color="bae5d1", fill_type="solid")
    rule1 = Rule(type="expression", dxf=DifferentialStyle(fill=green_fill))
    rule1.formula = ["$C$1*F3>(G3+H3)"]

    yellow_fill = PatternFill(start_color="FDDB6D", end_color="FDDB6D", fill_type="solid")
    rule3 = Rule(type="expression", dxf=DifferentialStyle(fill=yellow_fill))
    rule3.formula = ['W3<>""']
    sheet.conditional_formatting.add(('B3:B' + c), rule3)

    sheet.conditional_formatting.add(('F3:F' + c), rule1)

    sheet.freeze_panes = 'D3'

    sheet.sheet_view.zoomScale = 85


    wb.save(file_path + 'Заказ ' + prk_number + ' ' + now1 + '.xlsx')

    add_message('готово')
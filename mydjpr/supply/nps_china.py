import pandas as pd
import os
import glob
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from time import time
from python_calamine.pandas import pandas_monkeypatch

pandas_monkeypatch()

pd.options.mode.chained_assignment = None


os.chdir(os.path.dirname(os.path.abspath(__file__)))



def extract_nps(file_path, add_message):

    if not file_path.endswith(os.sep):
        file_path = file_path + os.sep
    t = time()
    d = {f: os.stat(f).st_mtime for f in glob.iglob(file_path + "NPS\\*.xlsx")}
    file_name_nps = max(d, key=lambda i: d[i])
    df = pd.read_excel(file_name_nps, sheet_name='общее', engine='calamine')
    print(df['Про доставку'])
    # Исключаем строки, где "Про доставку"==ЛОЖЬ
    df = df[df['Про доставку'] != 1].copy()
    print(df['Про доставку'])
    # Преобразуем столбец "Месяц и год" к дате
    df['_date'] = pd.to_datetime(df['Месяц и год'], dayfirst=True, errors='coerce')
    # Удалим строки без корректной даты
    df = df[~df['_date'].isna()].copy()
    # Рассчитываем границы периода: последние 9 месяцев, заканчивая предыдущим месяцем
    today = pd.Timestamp.today().normalize()
    prev_month_end = today.replace(day=1) - pd.offsets.Day(1)
    prev_month_start = prev_month_end.replace(day=1)
    def period_bounds(months):
        start = prev_month_start - pd.DateOffset(months=months-1)
        end = prev_month_end
        return start, end
    start9, end9 = period_bounds(9)
    df9 = df[(df['_date'] >= start9) & (df['_date'] <= end9)].copy()
    def build_stats(src):
        # Средняя оценка
        mean_rating = src.groupby('Код продукта')['Оценка'].mean().reset_index()
        mean_rating.columns = ['Код продукта', 'Средняя оценка']
        # Количество отзывов
        review_count = src.groupby('Код продукта')['Оценка'].count().reset_index()
        review_count.columns = ['Код продукта', 'Количество отзывов']
        # Объединение
        result = pd.merge(mean_rating, review_count, on='Код продукта')
        return result
    result_9m = build_stats(df9)
    # Сохраняем только лист с 9 месяцами
    output_path = 'nps_orders.xlsx'
    with pd.ExcelWriter(r'\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Шафиев\Обработка_заказ_Китай\nps_orders.xlsx') as writer:
        result_9m.to_excel(writer, index=False, float_format="%.2f", sheet_name='9 месяцев')
    print(f"Анализ завершен. Результат сохранен в: {output_path}")
    print(f"Период 9 месяцев: {start9.date()} — {end9.date()}")
    add_message('Файл обновлен!')

    


def normalize_code(x):
    """
    Приводит к float, потом к int (если возможно), потом к строке. 
    Если не конвертируется — к строке без пробелов.
    """
    try:
        return str(int(float(str(x).replace(',', '.').strip())))
    except Exception:
        return str(x).strip()

def add_to_orders(file_path, add_message):
    if not file_path.endswith(os.sep):
        file_path = file_path + os.sep
    # 1. Загрузка данных NPS (только 9 месяцев)
    nps_file = r'\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Шафиев\Обработка_заказ_Китай\nps_orders.xlsx'
    try:
        xls = pd.ExcelFile(nps_file, engine='calamine')
        has_9 = '9 месяцев' in xls.sheet_names
        if has_9:
            nps9 = pd.read_excel(nps_file, sheet_name='9 месяцев', engine='calamine')
        else:
            nps9 = pd.read_excel(nps_file, engine='calamine')  # fallback
            print('Лист "9 месяцев" не найден, читаю первый лист как 9 мес.')
    except Exception as e:
        print(f'Ошибка чтения файла NPS: {e}')
        add_message('Не удалось прочитать файл с NPS.')
        return
    
    # Приведение ключей к строке
    if 'Код продукта' not in nps9.columns:
        print('В данных NPS отсутствует столбец "Код продукта".')
        add_message('Некорректный формат файла NPS.')
        return
    nps9['Код продукта'] = nps9['Код продукта'].astype(str).str.strip()
    
    # Переименование столбцов для 9 месяцев
    nps9 = nps9.rename(columns={
        'Средняя оценка': 'Средняя оценка (9 мес)',
        'Количество отзывов': 'Количество отзывов (9 мес)',
    })
    
    # 1b. Загрузка Плюсы/Минусы (качество и доставка)
    pros_cons_file = r'\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Шафиев\Обработка_заказ_Китай\pros_cons.xlsx'
    pros_cons_quality, pros_cons_delivery = None, None
    if os.path.exists(pros_cons_file):
        def _load_proscons(sheet):
            try:
                df = pd.read_excel(pros_cons_file, sheet_name=sheet, engine='calamine')
                expected_cols = {'Код товара', 'Плюсы', 'Минусы'}
                real_cols = set(df.columns)
                if not expected_cols.issubset(real_cols):
                    print(f'В листе "{sheet}" отсутствуют столбцы: {", ".join(expected_cols - real_cols)}')
                    print('Столбцы в файле:', df.columns.tolist())
                    return None
                df['Код товара'] = df['Код товара'].apply(normalize_code)
                def _clean_val(x):
                    if pd.isna(x):
                        return ''
                    s = str(x).strip()
                    return '' if s in {'—', '-', '–'} else s
                df['Плюсы'] = df['Плюсы'].apply(_clean_val)
                df['Минусы'] = df['Минусы'].apply(_clean_val)
                return df
            except Exception as e:
                print(f'Ошибка чтения листа "{sheet}" из pros_cons: {e}')
                return None
        pros_cons_quality = _load_proscons('качество')
        pros_cons_delivery = _load_proscons('доставка')
        # DEBUG print
        print('pros_cons_quality sample:')
        if pros_cons_quality is not None:
            print(pros_cons_quality.head(3))
        else:
            print('pros_cons_quality NOT LOADED')
    else:
        print(f'Файл {os.path.basename(pros_cons_file)} не найден. Плюсы/минусы будут пустыми.')
    
    # 2. Обработка файлов Orders
    orders_folder = file_path + 'Orders'
    files = [f for f in os.listdir(orders_folder)
             if f.endswith('.xlsx') and f != os.path.basename(nps_file)]
    
    for file in files:
        print('ОБРАБОТКА:', file)
        file_path1 = os.path.join(orders_folder, file)
        # Определение типа файла
        first_row = pd.read_excel(file_path1, nrows=1, header=None).iloc[0, 0]
        file_type = 'type2' if "Период отчета:" in str(first_row) else 'type1'
        merge_col = ''
        header_row = 0
        
        if file_type == 'type1':
            df = pd.read_excel(file_path1, engine='calamine')
            art_column = next(col for col in df.columns if 'Артикул (доп)' in str(col))
            # Нормализация:
            df[art_column] = df[art_column].apply(normalize_code)
            merge_col = art_column
        else:
            for i in range(15, 20):
                df_test = pd.read_excel(file_path1, header=i, nrows=1, engine='calamine')
                if any('Код' in str(col) for col in df_test.columns):
                    header_row = i
                    break
            else:
                print(f"Не найден заголовок в файле {file}")
                continue
            df = pd.read_excel(file_path1, header=header_row, engine='calamine')
            code_column = next(col for col in df.columns if 'Код' in str(col))
            df[code_column] = df[code_column].apply(normalize_code)
            merge_col = code_column
        
        # DEBUG до merge
        print('merge_col:', merge_col)
        print('Первые коды из файла:', df[merge_col].unique()[:5])
        if pros_cons_quality is not None:
            print('Первые коды из pros_cons_quality:', pros_cons_quality['Код товара'].unique()[:5])

        # Мерджим NPS (только данные за 9 месяцев)
        nps9['Код продукта'] = nps9['Код продукта'].apply(normalize_code)
        nps_merged = df.merge(
            nps9[['Код продукта', 'Средняя оценка (9 мес)', 'Количество отзывов (9 мес)']],
            left_on=merge_col,
            right_on='Код продукта',
            how='left'
        )
        
        # Плюсы/Минусы качество
        if pros_cons_quality is not None:
            pm_q = df[[merge_col]].merge(
                pros_cons_quality[['Код товара', 'Плюсы', 'Минусы']],
                left_on=merge_col, right_on='Код товара', how='left'
            )
            print('После merge (качество):')
            print(pm_q.head(5))
            plus_q = pm_q['Плюсы'].fillna('').astype(str).tolist()
            minus_q = pm_q['Минусы'].fillna('').astype(str).tolist()
        else:
            plus_q, minus_q = [''] * len(df), [''] * len(df)
            print('pros_cons_quality = None, поэтому плюсы/минусы по качеству заполняются пустыми')
        
        # Плюсы/Минусы доставка
        if pros_cons_delivery is not None:
            pm_d = df[[merge_col]].merge(
                pros_cons_delivery[['Код товара', 'Плюсы', 'Минусы']],
                left_on=merge_col, right_on='Код товара', how='left'
            )
            print('После merge (доставка):')
            print(pm_d.head(5))
            plus_d = pm_d['Плюсы'].fillna('').astype(str).tolist()
            minus_d = pm_d['Минусы'].fillna('').astype(str).tolist()
        else:
            plus_d, minus_d = [''] * len(df), [''] * len(df)
        
        # ==== Запись ====
        wb = load_workbook(file_path1)
        ws = wb.active
        # Только столбцы за 9 месяцев и плюсы/минусы
        headers_to_add = [
            'Средняя оценка (9 мес)',
            'Количество отзывов (9 мес)',
            'Плюсы (качество)',
            'Минусы (качество)',
            'Плюсы (доставка)',
            'Минусы (доставка)'
        ]
        
        last_col = ws.max_column
        for offset, h in enumerate(headers_to_add, start=1):
            ws.cell(row=header_row+1, column=last_col+offset, value=h)
        
        # Заполняем строки
        for idx, (_, row) in enumerate(nps_merged.iterrows(), start=header_row+2):
            if idx > ws.max_row:
                break
            # Данные за 9 месяцев
            ws.cell(row=idx, column=last_col+1, value=row.get('Средняя оценка (9 мес)'))
            ws.cell(row=idx, column=last_col+2, value=row.get('Количество отзывов (9 мес)'))
            # Плюсы/минусы
            ws.cell(row=idx, column=last_col+3, value=plus_q[idx - (header_row + 2)])
            ws.cell(row=idx, column=last_col+4, value=minus_q[idx - (header_row + 2)])
            ws.cell(row=idx, column=last_col+5, value=plus_d[idx - (header_row + 2)])
            ws.cell(row=idx, column=last_col+6, value=minus_d[idx - (header_row + 2)])
        
        wb.save(file_path1)
        print(f"Обработан файл {file} (тип {file_type})")
    
    add_message('Готово! Добавлены данные за 9 месяцев (средняя оценка, количество отзывов) и плюсы/минусы по качеству и доставке.')

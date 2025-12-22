import pandas as pd
import numpy as np
import os
import glob
from datetime import datetime, timedelta
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill
import shutil
import time
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from pandas import read_excel
from python_calamine.pandas import pandas_monkeypatch


pandas_monkeypatch()



pd.options.mode.chained_assignment = None

os.chdir(os.path.dirname(os.path.abspath(__file__)))



def dop_shafiev(file_path, add_message):
    if not file_path.endswith(os.sep):
        file_path = file_path + os.sep
        
    d = {f: os.stat(f).st_mtime for f in glob.iglob(file_path + "Исходники\\*.csv")}
    file_name_csv = max(d, key=lambda i: d[i])
    # add_message(d)


    df_zakaz = pd.read_csv(file_name_csv, delimiter=';', low_memory=False, thousands=' ')
    

    # n = {f: os.stat(f).st_mtime for f in glob.iglob(r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Список_нел\\*.csv")}
    # file_name_nel = max(n, key=lambda i: n[i])
    # df_nel = pd.read_csv(file_name_nel, delimiter=';', low_memory=False)
    # df_nel['Неликвид'] = 'Да'
    # df_nel.rename(columns={'Код': 'Артикул(доп.)', 'Склад': 'В'}, inplace=True)


    n = {f: os.stat(f).st_mtime for f in glob.iglob(r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Список_нел\\*.parquet")}
    file_name_nel = max(n, key=lambda i: n[i])
    print(file_name_nel)
    df_nel = pd.read_parquet(file_name_nel)
    df_nel['Неликвид'] = 'Да'
    df_nel.rename(columns={'Код': 'Артикул(доп.)', 'Склад': 'В'}, inplace=True)

    # Выведенные позиции
    k = {f: os.stat(f).st_mtime for f in glob.iglob(r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Шафиев\Выведенные_коды_розница\\*.xlsx")}
    file_name_viv = max(k, key=lambda i: k[i])
    viv_column = 'Выведенный из розницы ' + datetime.fromtimestamp(k[file_name_viv]).strftime('%d-%m-%Y')
    # print(ost_rozn_column)
    df_viv = pd.read_excel(file_name_viv, engine="calamine")
    df_viv[viv_column] = 'Да'
    df_viv.rename(columns={'Код': 'Артикул(доп.)'}, inplace=True)
    


    d = {f: os.stat(f).st_mtime for f in glob.iglob(file_path + "Исходники\\*.xlsx")}
    files = []
    for i in d.keys():
        files.append(i)
    # d = {f: os.stat(f).st_mtime for f in glob.iglob("Состояние запасов 01.11.2023.xlsx")}
    # file_name_stel = max(d, key=lambda i: d[i])

    if 'Состояние' in files[0]:
        file_name_stel = files[0]
        shablon = files[1]
    else:
        file_name_stel = files[1]
        shablon = files[0]
    print(shablon)
    # add_message(shablon)

    df_prk = pd.read_excel(shablon)
    df_prk = df_prk.drop_duplicates(subset=['Артикул(доп.)'])
    prk_number = df_zakaz.iloc[1]['В']
    print(prk_number)
    add_message(prk_number)

    df1 = pd.read_excel(file_name_stel, engine="calamine")
    # df1.to_excel("123456.xlsx", index=False)
    i, j = np.where(df1.values == prk_number)
    print(i, j)
    add_message(i)
    add_message(j)
    


    column_prk_format = (int(j))
    column_prk_stel = (int(j+7))

    df_stel = df1[['Unnamed: 0', ('Unnamed: '+str(column_prk_stel)), ('Unnamed: '+str(column_prk_format))]]
    df_stel = df_stel.drop([0,1,2,3,4,5,6,7,8,9])

    df_stel = df_stel.rename(columns={'Unnamed: 0': 'СегментСтелажногоХранения',
                                    ('Unnamed: '+str(column_prk_stel)):'Постеллажка',
                                    ('Unnamed: ' + str(column_prk_format)): 'Формат'})


    df_wzakaz = df_zakaz.merge(df_prk, on='Артикул(доп.)', how='left')
    df_wstel = df_wzakaz.merge(df_stel, on='СегментСтелажногоХранения', how='left')
    df_wstel = df_wstel.drop_duplicates(subset='Артикул(доп.)', keep='first')

    # df_wstel['Остаток свободный (Из)'] = df_wstel['Остаток свободный (Из)'].astype(int)


    df_wstel['Остаток свободный (Из)'] = pd.to_numeric(df_wstel['Остаток свободный (Из)'], errors='coerce')
    df_wstel['Сумма продаж за 13 мес. (В)'] = pd.to_numeric(df_wstel['Сумма продаж за 13 мес. (В)'], errors='coerce')
    df_wstel['Минимальная партия,ед. (В)'] = pd.to_numeric(df_wstel['Минимальная партия,ед. (В)'], errors='coerce')

    df_nbn = df_wstel[df_wstel['Не балансировать товар (В)'] == 'Нет']
    df_nbn['Не > 2xОТЗ'] = (df_nbn['ОТЗ (В) на норму запаса'] * 2 - df_nbn['Остаток свободный (В)'] - df_nbn['Заказано (В)'])
    df_nbd = df_wstel.loc[df_wstel['Не балансировать товар (В)'] == 'Да']
    df_nbd['Не > 2xОТЗ'] = " "
    frames = [df_nbn, df_nbd]
    df_nb2 = pd.concat(frames)
    df_nb2['Постеллажка'].fillna(99999, inplace=True)


    df_nrb = df_nb2.loc[(df_nb2['Постеллажка'] != 99999) & (df_nb2['Не балансировать товар (В)'] == 'Да')]
    # df_nrb.to_excel("123.xlsx")
    df_nrb['Постеллажка'] = pd.to_numeric(df_nrb['Постеллажка'], errors='coerce')
    df_nrb = df_nrb[df_nrb['Постеллажка'] <= 0]
    df_nrb['Комментарий'] = "Прошу проверить постеллажку"

    df_em = df_nb2.loc[(df_nb2['Постеллажка'] != 99999) & (df_nb2['Не балансировать товар (В)'] == 'Да')]
    df_em['Постеллажка'] = pd.to_numeric(df_em['Постеллажка'], errors='coerce')
    df_em = df_em.loc[(df_em['Постеллажка'] > 0) & (df_em['Постеллажка'] < 99999)]
    df_em['Комментарий'] = " "

    df_nbrb = df_nb2.loc[(df_nb2['Не балансировать товар (В)'] == 'Нет') & (df_nb2['Постеллажка'] != 99999)]
    df_nbrb['Комментарий'] = " "

    # df_rzb = df_nb2.loc[(df_nb2['Постеллажка'] == 99999) & (df_nb2['Не балансировать товар (В)'] == 'Да')]
    # df_rzb['Комментарий'] = " "

    df_rb = df_nb2[df_nb2['Постеллажка'] == 99999]
    df_rb['Комментарий'] = "СЕГМЕНТ НЕ НАЗНАЧЕН"
    frames_2 = [df_nrb, df_em, df_nbrb, df_rb]
    df_nb3 = pd.concat(frames_2)
    df_nb3['Направления балансировки (В)'].fillna('не задано', inplace=True)

    df1 = df_nb3[df_nb3['Направления балансировки (В)'].str.contains("блок") == True]
    df1['Комментарий'] = 'Заблокировано'

    df2 = df_nb3[df_nb3['Направления балансировки (В)'].str.contains("блок") != True]
    frames_4 = [df1, df2]
    df3 = pd.concat(frames_4)

    df_nb4 = df3[df3['Остаток свободный (Из)'] == 0]
    df_nb4['Комментарий'] = "Нет на базе"

    df_nb5 = df3[df3['Остаток свободный (Из)'] != 0]
    frames_3 = [df_nb4, df_nb5]
    df_nb6 = pd.concat(frames_3)


    now = datetime.now()  - timedelta(days=1)
    now_month = now.strftime("%m-%Y")
    month1, year1 = (now.month-1, now.year) if now.month != 1 else (12, now.year-1)
    prev_month = now.replace(day=1, month=month1, year=year1).strftime("%m-%Y")
    month2, year2 = (now.month, now.year-1)
    prev_year_month = now.replace(day=1, month=month2, year=year2).strftime("%m-%Y")

    df_nb6['ЗАКАЗ'] = " "
    df_nb6['Сумма заказа'] = " "
    df_nb6['Объем заказа'] = " "
    df_nb6['Объем заказа1'] = " "

    # df_nb6['ЗАКАЗ'] = pd.to_numeric(df_nb6['ЗАКАЗ'], errors='coerce')
    df_nb6.loc[df_nb6['Постеллажка'] == 99999, 'Постеллажка'] = "нет значения"

    df_nb6 = df_nb6.sort_values(by=['СегментСтелажногоХранения', 'Наименование'])

    df_nb6['Можно выписать'] = df_nb6['ОТЗ (В) на норму запаса'] * 2 - (df_nb6['Остаток свободный (В)'] + df_nb6['Заказано (В)'])
    df_nb6.loc[df_nb6['Можно выписать'] < 0, 'Можно выписать'] = 0

    df_nb6.loc[df_nb6['Направления балансировки (В)'] == 'с РЦ на ПРК', 'ИТОГО выписать'] = df_nb6[['Можно выписать', 'Заказ ПРК', 'Остаток свободный (Из)']].min(axis=1)
    df_nb6.loc[df_nb6['Остаток свободный (Из)'] == 0, 'ИТОГО выписать'] = "нет на базе"
    df_nb6.loc[df_nb6['Направления балансировки (В)'].str.contains("блок") == True, 'ИТОГО выписать'] = "заблокировано"
    df_nb6.loc[df_nb6['Направления балансировки (В)'] == 'не задано', 'ИТОГО выписать'] = "неактив"
    

    df5 = df_nb6[['Артикул(доп.)', 'Наименование', 'В', 'Остаток свободный (Из)', 'Заказ ПРК', 'Можно выписать', 'ИТОГО выписать',
                    'Сумма заказа', 'Объем заказа',
                    'Комментарий', 'Постеллажка', 'Остаток свободный (В)', 'Заказано (В)', 'ОТЗ (В) на норму запаса',
                    'Суммарные продажи за 7 дней (В)', 'Продажи за '+ now_month +' (В)', 'Продажи за '+ prev_month +' (В)',
                    'Продажи за '+ prev_year_month +' (В)', 'Сумма продаж за 13 мес. (В)', 'КомментарийТО',
                    'Направления балансировки (В)', 'ЗакупочнаяЦена', 'Объем',
                    'СегментСтелажногоХранения', 'Минимальная партия,ед. (В)',
                    'Кратность ,ед. (В)', 'Не балансировать товар (В)', 'Каталог (В)',
                    'Маячки (В)', 'Формат', 'НСО_УстановленВДопЗонах']]
    
    df5 = df5.merge(df_nel, on=['Артикул(доп.)', 'В'], how='left')
    df5 = df5.merge(df_viv, on=['Артикул(доп.)'], how='left')
    df5.loc[df5['Неликвид'] == 'Да', 'ИТОГО выписать'] = "неликвид"
    df5.loc[df5[viv_column] == 'Да', 'ИТОГО выписать'] = "ВЫВЕДЕННЫЙ ИЗ ВСЕЙ РОЗНИЦЫ"
    

    now1 = datetime.now()
    now1 = now1.strftime("%d-%m-%Y")

    df5.to_excel(file_path + 'Заказ ' + prk_number + ' ' + now1 + '.xlsx', index=False)

    wb = openpyxl.load_workbook(file_path + 'Заказ ' + prk_number + ' ' + now1 + '.xlsx')
    sheet = wb.active
    sheet.insert_rows(0, amount=1)
    sheet.row_dimensions[2].height = 70
    sheet.column_dimensions['B'].width = 43
    for row in sheet.iter_rows(max_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    for i in range(2, (len(df5)+3)):
        sheet['G'+str(i)].fill = PatternFill(fill_type='solid', start_color='00FFFF99')

    for i in range(3, (len(df5)+3)):
        # sheet['H'+str(i)] = '=G%s*V%s' % (i, i)
        # sheet['I' + str(i)] = '=G%s*W%s' % (i, i)
        sheet[('H' + str(i))] = '=IF(ISNUMBER(G%s),V%s*G%s,0)' % (i, i, i)
        sheet[('I' + str(i))] = '=IF(ISNUMBER(G%s),W%s*G%s,0)' % (i, i, i)

    c = str(len(df5)+2)
    sheet['H1'] = '=SUM(H3:H' + c +')'
    sheet['I1'] = '=SUM(I3:I' + c +')'

    blue_fill = PatternFill(start_color="7cb5eb", end_color="7cb5eb", fill_type="solid")
    rule2 = Rule(type="expression", dxf=DifferentialStyle(fill=blue_fill))
    rule2.formula = ['AND(H3<200,H3<>0)']
    sheet.conditional_formatting.add(('H3:H' + c), rule2)
    
    pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
    rule2 = Rule(type="expression", dxf=DifferentialStyle(fill=pink_fill))
    rule2.formula = ['OR(K3<=0,K3="нет значения")']
    sheet.conditional_formatting.add(('K3:K' + c), rule2)

    rule3 = Rule(type="expression", dxf=DifferentialStyle(fill=pink_fill))
    rule3.formula = ['OR(AF3="Да", AG3="Да")']
    sheet.conditional_formatting.add(('B3:B' + c), rule3)

    
    sheet.sheet_view.zoomScale = 85


    wb.save(file_path + 'Заказ ' + prk_number + ' ' + now1 + '.xlsx')

    add_message('готово')



def dop_budyakova(file_path, add_message):
    if not file_path.endswith(os.sep):
        file_path = file_path + os.sep


    d = {f: os.stat(f).st_mtime for f in glob.iglob(file_path + "Исходники\\*.csv")}
    file_name_csv = max(d, key=lambda i: d[i])


    df_zakaz = pd.read_csv(file_name_csv, delimiter=';', low_memory=False, thousands=' ')

    n = {f: os.stat(f).st_mtime for f in glob.iglob(r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Список_нел\\*.parquet")}
    file_name_nel = max(n, key=lambda i: n[i])
    df_nel = pd.read_parquet(file_name_nel)
    df_nel['Неликвид'] = 'Да'
    df_nel.rename(columns={'Код': 'Артикул(доп.)', 'Склад': 'В'}, inplace=True)

    # Выведенные позиции
    k = {f: os.stat(f).st_mtime for f in glob.iglob(r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Шафиев\Выведенные_коды_розница\\*.xlsx")}
    file_name_viv = max(k, key=lambda i: k[i])
    viv_column = 'Выведенный из розницы ' + datetime.fromtimestamp(k[file_name_viv]).strftime('%d-%m-%Y')
    # print(ost_rozn_column)
    df_viv = pd.read_excel(file_name_viv, engine="calamine")
    df_viv[viv_column] = 'Да'
    df_viv.rename(columns={'Код': 'Артикул(доп.)'}, inplace=True)


    d = {f: os.stat(f).st_mtime for f in glob.iglob(file_path + "Исходники\\*.xlsx")}
    files = []
    for i in d.keys():
        files.append(i)


    if 'Состояние' in files[0]:
        file_name_stel = files[0]
        shablon = files[1]
    else:
        file_name_stel = files[1]
        shablon = files[0]
    # print(shablon)

    df_prk = pd.read_excel(shablon)
    prk_number = df_zakaz.iloc[0]['В']
    print(prk_number)
    add_message(prk_number)

    df1 = pd.read_excel(file_name_stel)
    i, j = np.where(df1.values == prk_number)
    print(i, j)
    column_prk_format = (int(j))
    column_prk_stel = (int(j+7))
    print("1")

    df_stel = df1[['Unnamed: 0', ('Unnamed: '+str(column_prk_stel)), ('Unnamed: '+str(column_prk_format))]]
    df_stel = df_stel.drop([0,1,2,3,4,5,6,7,8,9])

    df_stel = df_stel.rename(columns={'Unnamed: 0': 'СегментСтелажногоХранения',
                                    ('Unnamed: '+str(column_prk_stel)):'Постеллажка',
                                    ('Unnamed: ' + str(column_prk_format)): 'Формат'})

    df_wzakaz = df_zakaz.merge(df_prk, on='Артикул(доп.)', how='left')
    df_wstel = df_wzakaz.merge(df_stel, on='СегментСтелажногоХранения', how='left')
    df_wstel = df_wstel.drop_duplicates(subset='Артикул(доп.)', keep='first')

    # df_wstel['Остаток свободный (Из)'] = df_wstel['Остаток свободный (Из)'].astype(int)


    df_wstel['Остаток свободный (Из)'] = pd.to_numeric(df_wstel['Остаток свободный (Из)'], errors='coerce')
    df_wstel['Сумма продаж за 13 мес. (В)'] = pd.to_numeric(df_wstel['Сумма продаж за 13 мес. (В)'], errors='coerce')
    df_wstel['Минимальная партия,ед. (В)'] = pd.to_numeric(df_wstel['Минимальная партия,ед. (В)'], errors='coerce')

    df_nbn = df_wstel[df_wstel['Не балансировать товар (В)'] == 'Нет']
    df_nbn['Не > 2xОТЗ'] = (df_nbn['ОТЗ (В) на норму запаса'] * 2 - df_nbn['Остаток свободный (В)'] - df_nbn['Заказано (В)'])
    df_nbd = df_wstel.loc[df_wstel['Не балансировать товар (В)'] == 'Да']
    df_nbd['Не > 2xОТЗ'] = " "
    frames = [df_nbn, df_nbd]
    df_nb2 = pd.concat(frames)
    df_nb2['Постеллажка'].fillna(99999, inplace=True)

    now = datetime.now() - timedelta(days=1)
    now_month = now.strftime("%m-%Y")
    month1, year1 = (now.month-1, now.year) if now.month != 1 else (12, now.year-1)
    prev_month = now.replace(day=1, month=month1, year=year1).strftime("%m-%Y")
    month2, year2 = (now.month, now.year-1)
    prev_year_month = now.replace(day=1, month=month2, year=year2).strftime("%m-%Y")
    month3, year3 = (now.month+1, now.year-1) if now.month != 12 else (1, now.year)
    prev_year_month2 = now.replace(day=1, month=month3, year=year3).strftime("%m-%Y")

    df_nb2['ЗАКАЗ'] = " "
    df_nb2['Сумма заказа'] = " "
    df_nb2['Объем заказа'] = " "
    df_nb2['Можно выписать'] = " "

    df_nb2['ЗАКАЗ'] = pd.to_numeric(df_nb2['ЗАКАЗ'], errors='coerce')

    df_nb2 = df_nb2.sort_values(by=['СегментСтелажногоХранения', 'Наименование'])
    print(prev_year_month, prev_year_month2)
    df_nb2.loc[df_nb2['Постеллажка'] == 99999, 'Постеллажка'] = "нет значения"

    df5 = df_nb2[['Артикул(доп.)', 'Наименование', 'В', 'Заказ ПРК', 'Можно выписать', 'Количество для перемещения', 'ОТЗ (В) на норму запаса',
                    'Остаток свободный (В)', 'Заказано (В)', 'Сумма заказа', 'Остаток свободный (Из)', 'Объем заказа',
                    'ЗакупочнаяЦена', 'КаналПоставки', 'СегментСтелажногоХранения', 'Сезон',
                'Направления балансировки (В)', 'Роль ассортимента (В)', 'Кратность ,ед. (В)', 'Минимальный запас, ед. (В)',
                'Каталог (В)', 'Маячки (В)', 'Период действия будущей акции (В)',
                'Период действия текущей акции (В)',
                    'Суммарные продажи за 7 дней (В)',
                    'Продажи за '+ prev_year_month +' (В)', 'Продажи за '+ prev_year_month2 +' (В)', 'Объем',
                    'Не балансировать товар (В)', 'Постеллажка', 'Формат', 'НСО_УстановленВДопЗонах']]
    
    df5 = df5.merge(df_nel, on=['Артикул(доп.)', 'В'], how='left')
    df5 = df5.merge(df_viv, on=['Артикул(доп.)'], how='left')
    df5.loc[df5['Неликвид'] == 'Да', 'Количество для перемещения'] = "неликвид"
    df5.loc[df5[viv_column] == 'Да', 'Количество для перемещения'] = "ВЫВЕДЕННЫЙ ИЗ ВСЕЙ РОЗНИЦЫ"

    now1 = datetime.now()
    now1 = now1.strftime("%d-%m-%Y")

    df5.to_excel(file_path + 'Заказ ' + prk_number + ' ' + now1 + '.xlsx', index=False)

    wb = openpyxl.load_workbook(file_path + 'Заказ ' + prk_number + ' ' + now1 + '.xlsx')
    sheet = wb.active
    sheet.insert_rows(0, amount=1)
    sheet.row_dimensions[2].height = 90
    sheet.column_dimensions['B'].width = 43
    sheet.column_dimensions['C'].width = 13
    for row in sheet.iter_rows(max_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    for col in range(openpyxl.utils.column_index_from_string('G'), openpyxl.utils.column_index_from_string('L') + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 7

    for col in range(openpyxl.utils.column_index_from_string('Q'), openpyxl.utils.column_index_from_string('AA') + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 7

    for i in range(2, (len(df5)+3)):
        sheet['F'+str(i)].fill = PatternFill(fill_type='solid', start_color='00FFFF99')

    for i in range(3, (len(df5)+3)):
        # sheet['J'+str(i)] = '=M%s*F%s' % (i, i)
        # sheet['L' + str(i)] = '=AB%s*F%s' % (i, i)
        sheet[('J' + str(i))] = '=IF(ISNUMBER(F%s),M%s*F%s,0)' % (i, i, i)
        sheet[('L' + str(i))] = '=IF(ISNUMBER(F%s),AB%s*F%s,0)' % (i, i, i)
        sheet['E' + str(i)] = '=G%s*2-I%s-H%s' % (i, i, i)

    c = str(len(df5)+2)
    sheet['J1'] = '=SUM(J3:J' + c +')'
    sheet['L1'] = '=SUM(L3:L' + c +')'


    blue_fill = PatternFill(start_color="7cb5eb", end_color="7cb5eb", fill_type="solid")
    rule2 = Rule(type="expression", dxf=DifferentialStyle(fill=blue_fill))
    rule2.formula = ['AND(J3<200,J3<>0)']
    sheet.conditional_formatting.add(('J3:J' + c), rule2)
    
    pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
    rule2 = Rule(type="expression", dxf=DifferentialStyle(fill=pink_fill))
    rule2.formula = ['OR(AD3<=0,AD3="нет значения")']
    sheet.conditional_formatting.add(('AD3:AD' + c), rule2)

    rule3 = Rule(type="expression", dxf=DifferentialStyle(fill=pink_fill))
    rule3.formula = ['OR(AG3="Да", AH3="Да")']
    sheet.conditional_formatting.add(('B3:B' + c), rule3)


    wb.save(file_path + 'Заказ ' + prk_number + ' ' + now1 + '.xlsx')

    add_message('Готово!')



def dop_grechushkin(file_path, add_message):
    if not file_path.endswith(os.sep):
        file_path = file_path + os.sep

    d = {f: os.stat(f).st_mtime for f in glob.iglob(file_path + "Исходники\\*.csv")}
    file_name_csv = max(d, key=lambda i: d[i])


    df_zakaz = pd.read_csv(file_name_csv, delimiter=';', low_memory=False, thousands=' ')

    n = {f: os.stat(f).st_mtime for f in glob.iglob(r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Список_нел\\*.parquet")}
    file_name_nel = max(n, key=lambda i: n[i])
    df_nel = pd.read_parquet(file_name_nel)
    df_nel['Неликвид'] = 'Да'
    df_nel.rename(columns={'Код': 'Артикул(доп.)', 'Склад': 'В'}, inplace=True)

    # Выведенные позиции
    k = {f: os.stat(f).st_mtime for f in glob.iglob(r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Шафиев\Выведенные_коды_розница\\*.xlsx")}
    file_name_viv = max(k, key=lambda i: k[i])
    viv_column = 'Выведенный из розницы ' + datetime.fromtimestamp(k[file_name_viv]).strftime('%d-%m-%Y')
    # print(ost_rozn_column)
    df_viv = pd.read_excel(file_name_viv, engine="calamine")
    df_viv[viv_column] = 'Да'
    df_viv.rename(columns={'Код': 'Артикул(доп.)'}, inplace=True)


    d = {f: os.stat(f).st_mtime for f in glob.iglob(file_path + "Исходники\\*.xlsx")}
    files = []
    for i in d.keys():
        files.append(i)


    if 'Состояние' in files[0]:
        file_name_stel = files[0]
        shablon = files[1]
    else:
        file_name_stel = files[1]
        shablon = files[0]
    print(shablon)

    df_prk = pd.read_excel(shablon)
    prk_number = df_zakaz.iloc[1]['В']
    print(prk_number)
    add_message(prk_number)

    df1 = pd.read_excel(file_name_stel)
    i, j = np.where(df1.values == prk_number)
    print(i, j)
    column_prk_format = (int(j))
    column_prk_stel = (int(j+7))

    df_stel = df1[['Unnamed: 0', ('Unnamed: '+str(column_prk_stel)), ('Unnamed: '+str(column_prk_format))]]
    df_stel = df_stel.drop([0,1,2,3,4,5,6,7,8,9])

    df_stel = df_stel.rename(columns={'Unnamed: 0': 'СегментСтелажногоХранения',
                                    ('Unnamed: '+str(column_prk_stel)):'Постеллажка',
                                    ('Unnamed: ' + str(column_prk_format)): 'Формат'})

    df_wzakaz = df_zakaz.merge(df_prk, on='Артикул(доп.)', how='left')
    df_wstel = df_wzakaz.merge(df_stel, on='СегментСтелажногоХранения', how='left')
    df_wstel = df_wstel.drop_duplicates(subset='Артикул(доп.)', keep='first')

    # df_wstel['Остаток свободный (Из)'] = df_wstel['Остаток свободный (Из)'].astype(int)


    df_wstel['Остаток свободный (Из)'] = pd.to_numeric(df_wstel['Остаток свободный (Из)'], errors='coerce')
    # df_wstel['Сумма продаж за 13 мес. (В)'] = pd.to_numeric(df_wstel['Сумма продаж за 13 мес. (В)'], errors='coerce')
    df_wstel['Минимальная партия,ед. (В)'] = pd.to_numeric(df_wstel['Минимальная партия,ед. (В)'], errors='coerce')

    df_nbn = df_wstel[df_wstel['Не балансировать товар (В)'] == 'Нет']
    df_nbn['Не > 2xОТЗ'] = (df_nbn['ОТЗ (В) на норму запаса'] * 2 - df_nbn['Остаток свободный (В)'] - df_nbn['Заказано (В)'])
    df_nbd = df_wstel.loc[df_wstel['Не балансировать товар (В)'] == 'Да']
    df_nbd['Не > 2xОТЗ'] = " "
    frames = [df_nbn, df_nbd]
    df_nb2 = pd.concat(frames)
    df_nb2['Постеллажка'].fillna(99999, inplace=True)

    now = datetime.now() - timedelta(days=1)
    now_month = now.strftime("%m-%Y")
    month1, year1 = (now.month-1, now.year) if now.month != 1 else (12, now.year-1)
    prev_month = now.replace(day=1, month=month1, year=year1).strftime("%m-%Y")
    month2, year2 = (now.month, now.year-1)
    prev_year_month = now.replace(day=1, month=month2, year=year2).strftime("%m-%Y")
    month3, year3 = (now.month+1, now.year-1) if now.month != 12 else (1, now.year)
    prev_year_month2 = now.replace(day=1, month=month3, year=year3).strftime("%m-%Y")

    df_nb2['ЗАКАЗ'] = " "
    df_nb2['Сумма заказа'] = " "
    df_nb2['Объем заказа'] = " "
    df_nb2['Расчетное значение'] = " "

    df_nb2['ЗАКАЗ'] = pd.to_numeric(df_nb2['ЗАКАЗ'], errors='coerce')

    df_nb2 = df_nb2.sort_values(by=['СегментСтелажногоХранения', 'Наименование'])
    # print(prev_year_month, prev_year_month2)

    df5 = df_nb2[['Артикул(доп.)', 'Наименование', 'В', 'Остаток свободный (Из)', 'Заказ ПРК', 'Количество для перемещения',
                'Расчетное значение', 'ОТЗ (В) на норму запаса',
                    'Остаток свободный (В)', 'Заказано (В)', 'Сумма заказа', 'Объем заказа',
                    'ЗакупочнаяЦена',  'Суммарные продажи за 7 дней (В)',
                    'Продажи за '+ prev_year_month +' (В)', 'Продажи за '+ prev_year_month2 +' (В)',
                'КаналПоставки',
                    'Не балансировать товар (В)', 'Постеллажка', 'СегментСтелажногоХранения', 'Сезон',
                'Направления балансировки (В)', 'Кратность ,ед. (В)', 'Минимальный запас, ед. (В)',
                'Каталог (В)', 'Маячки (В)', 'Период действия будущей акции (В)',
                'Период действия текущей акции (В)', 'Минимальная партия,ед. (В)', 'Объем',
                'Формат', 'НСО_УстановленВДопЗонах']]
    
    df5 = df5.merge(df_nel, on=['Артикул(доп.)', 'В'], how='left')
    df5 = df5.merge(df_viv, on=['Артикул(доп.)'], how='left')
    df5.loc[df5['Неликвид'] == 'Да', 'Количество для перемещения'] = "неликвид"
    df5.loc[df5[viv_column] == 'Да', 'Количество для перемещения'] = "ВЫВЕДЕННЫЙ ИЗ ВСЕЙ РОЗНИЦЫ"

    now1 = datetime.now()
    now1 = now1.strftime("%d-%m-%Y")

    df5.to_excel(file_path + 'Заказ ' + prk_number + ' ' + now1 + '.xlsx', index=False)

    wb = openpyxl.load_workbook(file_path + 'Заказ ' + prk_number + ' ' + now1 + '.xlsx')
    sheet = wb.active
    sheet.insert_rows(0, amount=1)
    sheet.row_dimensions[2].height = 90
    sheet.column_dimensions['B'].width = 43
    sheet.column_dimensions['C'].width = 13
    for row in sheet.iter_rows(max_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    for col in range(openpyxl.utils.column_index_from_string('G'), openpyxl.utils.column_index_from_string('L') + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 7

    for col in range(openpyxl.utils.column_index_from_string('Q'), openpyxl.utils.column_index_from_string('AA') + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 7

    for i in range(2, (len(df5)+3)):
        sheet['E'+str(i)].fill = PatternFill(fill_type='solid', start_color='00FFFF99')
        sheet['F' + str(i)].fill = PatternFill(fill_type='solid', start_color='00FFFF99')

    for i in range(3, (len(df5)+3)):
        # sheet['K'+str(i)] = '=M%s*F%s' % (i, i)
        # sheet['L' + str(i)] = '=AD%s*F%s' % (i, i)
        sheet['G' + str(i)] = '=H%s*2-I%s-J%s' % (i, i, i)
        sheet[('K' + str(i))] = '=IF(ISNUMBER(F%s),M%s*F%s,0)' % (i, i, i)
        sheet[('L' + str(i))] = '=IF(ISNUMBER(F%s),AD%s*F%s,0)' % (i, i, i)

    c = str(len(df5)+2)
    sheet['L1'] = '=SUM(L3:L' + c +')'
    sheet['K1'] = '=SUM(K3:K' + c +')'

    pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
    rule2 = Rule(type="expression", dxf=DifferentialStyle(fill=pink_fill))
    rule2.formula = ["S3<=0"]
    sheet.conditional_formatting.add(('S3:S' + c), rule2)

    rule = Rule(type="expression", dxf=DifferentialStyle(fill=pink_fill))
    rule.formula = ['R3="Да"']
    sheet.conditional_formatting.add(('R3:R' + c), rule)

    rule1 = Rule(type="expression", dxf=DifferentialStyle(fill=pink_fill))
    rule1.formula = ['AND(F3*M3<200,F3<>0)']
    sheet.conditional_formatting.add(('F3:F' + c), rule1)

    rule3 = Rule(type="expression", dxf=DifferentialStyle(fill=pink_fill))
    rule3.formula = ['OR(AG3="Да", AH3="Да")']
    sheet.conditional_formatting.add(('B3:B' + c), rule3)

    filename = 'Заказ ' + prk_number + ' ' + now1 + '.xlsx'

    wb.save(file_path + filename)

    source_file_path = file_path + filename

    destination_folder =r'\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\В.Гречушкин\Общая\Доп. заказы'

    destination_file_path = os.path.join(destination_folder, filename)
    shutil.copy2(source_file_path, destination_file_path)

    add_message('Готово!')


def dop_kunavina(file_path, add_message):
    if not file_path.endswith(os.sep):
        file_path = file_path + os.sep

    d = {f: os.stat(f).st_mtime for f in glob.iglob(file_path + "Исходники\\*.csv")}
    file_name_csv = max(d, key=lambda i: d[i])


    df_zakaz = pd.read_csv(file_name_csv, delimiter=';', low_memory=False, thousands=' ')

    n = {f: os.stat(f).st_mtime for f in glob.iglob(r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Список_нел\\*.parquet")}
    file_name_nel = max(n, key=lambda i: n[i])
    df_nel = pd.read_parquet(file_name_nel)
    df_nel['Неликвид'] = 'Да'
    df_nel.rename(columns={'Код': 'Артикул(доп.)', 'Склад': 'В'}, inplace=True)

    # Выведенные позиции
    k = {f: os.stat(f).st_mtime for f in glob.iglob(r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Шафиев\Выведенные_коды_розница\\*.xlsx")}
    file_name_viv = max(k, key=lambda i: k[i])
    viv_column = 'Выведенный из розницы ' + datetime.fromtimestamp(k[file_name_viv]).strftime('%d-%m-%Y')
    # print(ost_rozn_column)
    df_viv = pd.read_excel(file_name_viv, engine="calamine")
    df_viv[viv_column] = 'Да'
    df_viv.rename(columns={'Код': 'Артикул(доп.)'}, inplace=True)


    d = {f: os.stat(f).st_mtime for f in glob.iglob(file_path + "Исходники\\*.xlsx")}
    files = []
    for i in d.keys():
        files.append(i)


    if 'Состояние' in files[0]:
        file_name_stel = files[0]
        shablon = files[1]
    else:
        file_name_stel = files[1]
        shablon = files[0]
    print(shablon)

    df_prk = pd.read_excel(shablon)
    prk_number = df_zakaz.iloc[1]['В']
    print(prk_number)
    add_message(prk_number)

    df1 = pd.read_excel(file_name_stel)
    i, j = np.where(df1.values == prk_number)
    print(i, j)
    column_prk_format = (int(j))
    column_prk_stel = (int(j+7))

    df_stel = df1[['Unnamed: 0', ('Unnamed: '+str(column_prk_stel)), ('Unnamed: '+str(column_prk_format))]]
    df_stel = df_stel.drop([0,1,2,3,4,5,6,7,8,9])

    df_stel = df_stel.rename(columns={'Unnamed: 0': 'СегментСтелажногоХранения',
                                    ('Unnamed: '+str(column_prk_stel)):'Постеллажка',
                                    ('Unnamed: ' + str(column_prk_format)): 'Формат'})

    df_wzakaz = df_zakaz.merge(df_prk, on='Артикул(доп.)', how='left')
    df_wstel = df_wzakaz.merge(df_stel, on='СегментСтелажногоХранения', how='left')
    df_wstel = df_wstel.drop_duplicates(subset='Артикул(доп.)', keep='first')

    # df_wstel['Остаток свободный (Из)'] = df_wstel['Остаток свободный (Из)'].astype(int)


    df_wstel['Остаток свободный (Из)'] = pd.to_numeric(df_wstel['Остаток свободный (Из)'], errors='coerce')
    df_wstel['Сумма продаж за 25 мес. (В)'] = pd.to_numeric(df_wstel['Сумма продаж за 25 мес. (В)'], errors='coerce')
    df_wstel['Минимальная партия,ед. (В)'] = pd.to_numeric(df_wstel['Минимальная партия,ед. (В)'], errors='coerce')

    df_nbn = df_wstel[df_wstel['Не балансировать товар (В)'] == 'Нет']
    df_nbn['Не > 2xОТЗ'] = (df_nbn['ОТЗ (В) на норму запаса'] * 2 - df_nbn['Остаток свободный (В)'] - df_nbn['Заказано (В)'])
    df_nbd = df_wstel.loc[df_wstel['Не балансировать товар (В)'] == 'Да']
    df_nbd['Не > 2xОТЗ'] = " "
    frames = [df_nbn, df_nbd]
    df_nb2 = pd.concat(frames)
    df_nb2['Постеллажка'].fillna(99999, inplace=True)

    now = datetime.now() - timedelta(days=1)
    now_month = now.strftime("%m-%Y")
    month1, year1 = (now.month-1, now.year) if now.month != 1 else (12, now.year-1)
    prev_month = now.replace(day=1, month=month1, year=year1).strftime("%m-%Y")
    month2, year2 = (now.month, now.year-1)
    prev_year_month = now.replace(day=1, month=month2, year=year2).strftime("%m-%Y")
    month3, year3 = (now.month+1, now.year-1) if now.month != 12 else (1, now.year)
    prev_year_month2 = now.replace(day=1, month=month3, year=year3).strftime("%m-%Y")

    df_nb2['ЗАКАЗ'] = " "
    df_nb2['Сумма заказа'] = " "
    df_nb2['Объем заказа'] = " "

    df_nb2['ЗАКАЗ'] = pd.to_numeric(df_nb2['ЗАКАЗ'], errors='coerce')

    df_nb2 = df_nb2.sort_values(by=['СегментСтелажногоХранения', 'Наименование'])
    print(prev_year_month, prev_year_month2)

    df5 = df_nb2[['Артикул(доп.)', 'Наименование', 'В', 'Заказ ПРК', 'Количество для перемещения', 'ОТЗ (В) на норму запаса',
                    'Остаток свободный (В)', 'Заказано (В)', 'Сумма заказа', 'Остаток свободный (Из)', 'Объем заказа',
                    'ЗакупочнаяЦена', 'КаналПоставки', 'СегментСтелажногоХранения', 'Сезон',
                'Направления балансировки (В)', 'Роль ассортимента (В)', 'Кратность ,ед. (В)', 'Минимальный запас, ед. (В)',
                'Каталог (В)', 'Маячки (В)', 'Период действия будущей акции (В)',
                'Период действия текущей акции (В)',
                    'Суммарные продажи за 7 дней (В)',
                    'Продажи за '+ prev_year_month +' (В)', 'Продажи за '+ prev_year_month2 +' (В)', 'Объем',
                    'Не балансировать товар (В)', 'Постеллажка', 'Формат', 'НСО_УстановленВДопЗонах']]
    
    df5 = df5.merge(df_nel, on=['Артикул(доп.)', 'В'], how='left')
    df5 = df5.merge(df_viv, on=['Артикул(доп.)'], how='left')
    df5.loc[df5['Неликвид'] == 'Да', 'Количество для перемещения'] = "неликвид"
    df5.loc[df5[viv_column] == 'Да', 'Количество для перемещения'] = "ВЫВЕДЕННЫЙ ИЗ ВСЕЙ РОЗНИЦЫ"

    now1 = datetime.now()
    now1 = now1.strftime("%d-%m-%Y")

    df5.to_excel(file_path + 'Заказ ' + prk_number + ' ' + now1 + '.xlsx', index=False)

    wb = openpyxl.load_workbook(file_path + 'Заказ ' + prk_number + ' ' + now1 + '.xlsx')
    sheet = wb.active
    sheet.insert_rows(0, amount=1)
    sheet.row_dimensions[2].height = 90
    sheet.column_dimensions['B'].width = 43
    sheet.column_dimensions['C'].width = 13
    for row in sheet.iter_rows(max_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    for col in range(openpyxl.utils.column_index_from_string('G'), openpyxl.utils.column_index_from_string('L') + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 7

    for col in range(openpyxl.utils.column_index_from_string('Q'), openpyxl.utils.column_index_from_string('AA') + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 7

    for i in range(2, (len(df5)+3)):
        sheet['E'+str(i)].fill = PatternFill(fill_type='solid', start_color='00FFFF99')

    for i in range(3, (len(df5)+3)):
        # sheet['I'+str(i)] = '=L%s*E%s' % (i, i)
        # sheet['K' + str(i)] = '=Z%s*E%s' % (i, i)
        sheet[('I' + str(i))] = '=IF(ISNUMBER(E%s),L%s*E%s,0)' % (i, i, i)
        sheet[('K' + str(i))] = '=IF(ISNUMBER(E%s),Z%s*E%s,0)' % (i, i, i)

    c = str(len(df5)+2)
    

    pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
    rule3 = Rule(type="expression", dxf=DifferentialStyle(fill=pink_fill))
    rule3.formula = ['OR(AF3="Да", AG3="Да")']
    sheet.conditional_formatting.add(('B3:B' + c), rule3)

    sheet['I1'] = '=SUM(I3:I' + c +')'
    sheet['K1'] = '=SUM(K3:K' + c +')'


    wb.save(file_path + 'Заказ ' + prk_number + ' ' + now1 + '.xlsx')

    add_message('Готово!')


def dop_torgashina(file_path, add_message):
    if not file_path.endswith(os.sep):
        file_path = file_path + os.sep

    d = {f: os.stat(f).st_mtime for f in glob.iglob(file_path + "Исходники\\*.csv")}
    file_name_csv = max(d, key=lambda i: d[i])


    df_zakaz = pd.read_csv(file_name_csv, delimiter=';', low_memory=False, thousands=' ')

    n = {f: os.stat(f).st_mtime for f in glob.iglob(r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Список_нел\\*.parquet")}
    file_name_nel = max(n, key=lambda i: n[i])
    df_nel = pd.read_parquet(file_name_nel)
    df_nel['Неликвид'] = 'Да'
    df_nel.rename(columns={'Код': 'Артикул(доп.)', 'Склад': 'В'}, inplace=True)

    # Выведенные позиции
    k = {f: os.stat(f).st_mtime for f in glob.iglob(r"\\lan.sct.ru\x\Воронеж\Подразделения\Коммерческий\Т.Шафиев\Общая\supply\Шафиев\Выведенные_коды_розница\\*.xlsx")}
    file_name_viv = max(k, key=lambda i: k[i])
    viv_column = 'Выведенный из розницы ' + datetime.fromtimestamp(k[file_name_viv]).strftime('%d-%m-%Y')
    # print(ost_rozn_column)
    df_viv = pd.read_excel(file_name_viv, engine="calamine")
    df_viv[viv_column] = 'Да'
    df_viv.rename(columns={'Код': 'Артикул(доп.)'}, inplace=True)


    d = {f: os.stat(f).st_mtime for f in glob.iglob(file_path + "Исходники\\*.xlsx")}
    files = []
    for i in d.keys():
        files.append(i)


    if 'Состояние' in files[0]:
        file_name_stel = files[0]
        shablon = files[1]
    else:
        file_name_stel = files[1]
        shablon = files[0]
    print(shablon)

    df_prk = pd.read_excel(shablon)
    prk_number = df_zakaz.iloc[1]['В']
    print(prk_number)
    add_message(prk_number)

    df1 = pd.read_excel(file_name_stel)
    i, j = np.where(df1.values == prk_number)
    print(i, j)
    column_prk_format = (int(j))
    column_prk_stel = (int(j+7))

    df_stel = df1[['Unnamed: 0', ('Unnamed: '+str(column_prk_stel)), ('Unnamed: '+str(column_prk_format))]]
    df_stel = df_stel.drop([0,1,2,3,4,5,6,7,8,9])

    df_stel = df_stel.rename(columns={'Unnamed: 0': 'СегментСтелажногоХранения',
                                  ('Unnamed: '+str(column_prk_stel)):'Блокировка',
                                  ('Unnamed: ' + str(column_prk_format)): 'Формат'})
    
    df_wzakaz = df_zakaz.merge(df_prk, on='Артикул(доп.)', how='left')
    df_wstel = df_wzakaz.merge(df_stel, on='СегментСтелажногоХранения', how='left')
    df_wstel = df_wstel.drop_duplicates(subset='Артикул(доп.)', keep='first')

    df_wstel['Остаток свободный (Из)'] = pd.to_numeric(df_wstel['Остаток свободный (Из)'], errors='coerce')
    df_wstel['Направления балансировки (В)'].fillna('не задано', inplace=True)
    df_wstel['Блокировка'].fillna('нет значения', inplace=True)

    now = datetime.now()
    now_month = now.strftime("%m-%Y")

    now_day = int(datetime.now().strftime("%d"))
    if now_day in range(1, 16):
        month2, year2 = (now.month, now.year - 1)
        prev_year_month = now.replace(day=1, month=month2, year=year2).strftime("%m-%Y")
    else:
        month3, year3 = (now.month + 1, now.year - 1) if now.month != 12 else (1, now.year)
        prev_year_month = now.replace(day=1, month=month3, year=year3).strftime("%m-%Y")

    df_wstel['Сумма заказа'] = " "
    df_wstel['Комментарий'] = " "
    df_wstel['Можно выписать'] = df_wstel['ОТЗ (В) на норму запаса'] * 2 - (df_wstel['Остаток свободный (В)'] + df_wstel['Заказано (В)'])
    df_wstel.loc[df_wstel['Можно выписать'] < 0, 'Можно выписать'] = 0
    df_wstel['Остаток+заказано в ПРК'] = df_wstel['Остаток свободный (В)'] + df_wstel['Заказано (В)']
    df_wstel['ОТЗ*2'] = df_wstel['ОТЗ (В) на норму запаса'] * 2

    df_wstel.loc[df_wstel['Направления балансировки (В)'] == 'с РЦ на ПРК', 'ИТОГО выписать'] = df_wstel[['Можно выписать', 'Заказ ПРК', 'Остаток свободный (Из)']].min(axis=1)
    df_wstel.loc[df_wstel['Остаток свободный (Из)'] == 0, 'ИТОГО выписать'] = "нет на базе"
    df_wstel.loc[df_wstel['Направления балансировки (В)'].str.contains("блок") == True, 'ИТОГО выписать'] = "заблокировано"
    df_wstel.loc[df_wstel['Направления балансировки (В)'] == 'не задано', 'ИТОГО выписать'] = "неактив"

    df_wstel = df_wstel.merge(df_nel, on=['Артикул(доп.)', 'В'], how='left')
    df_wstel.loc[df_wstel['Неликвид'] == 'Да', 'ИТОГО выписать'] = "неликвид"
    df_wstel = df_wstel.merge(df_viv, on=['Артикул(доп.)'], how='left')
    df_wstel.loc[df_wstel[viv_column] == 'Да', 'ИТОГО выписать'] = "ВЫВЕДЕННЫЙ ИЗ ВСЕЙ РОЗНИЦЫ"
    




    df1 = df_wstel[['Артикул(доп.)', 'Наименование', 'ЗакупочнаяЦена', 'Остаток свободный (Из)', 'Заказ ПРК',
                    'ИТОГО выписать', 'Сумма заказа', 'Комментарий', 'Можно выписать', 'Остаток свободный (В)',
                    'Заказано (В)', 'Остаток+заказано в ПРК', 'ОТЗ*2', 'Суммарные продажи за 7 дней (В)',
                    'Продажи за '+ prev_year_month +' (В)',
                    'Кратность ,ед. (В)', 'СегментСтелажногоХранения', 'Блокировка', 'Направления балансировки (В)',
                    'Не балансировать товар (В)', 'Каталог (В)', 'Маячки (В)', 'Период действия будущей акции (В)',
                    'Период действия текущей акции (В)', 'Неликвид', viv_column]]
    
    

    df1 = df1.sort_values(by=['СегментСтелажногоХранения', 'Наименование'])

    now1 = datetime.now()
    now1 = now1.strftime("%d-%m-%Y")

    df1.to_excel(file_path + 'Заказ ' + prk_number + ' ' + now1 + '.xlsx', index=False)


    wb = openpyxl.load_workbook(file_path + 'Заказ ' + prk_number + ' ' + now1 + '.xlsx')
    sheet = wb.active
    sheet.insert_rows(0, amount=1)
    sheet.row_dimensions[2].height = 90
    sheet.column_dimensions['B'].width = 49
    sheet.column_dimensions['Q'].width = 20
    sheet.column_dimensions['S'].width = 17
    for row in sheet.iter_rows(max_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    for col in range(openpyxl.utils.column_index_from_string('J'), openpyxl.utils.column_index_from_string('O') + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 7

    for i in range(2, (len(df1)+3)):
        sheet['F'+str(i)].fill = PatternFill(fill_type='solid', start_color='00FFFF99')

    for i in range(3, len(df1) + 3):
        sheet[('G' + str(i))] = '=IF(ISNUMBER(F%s),C%s*F%s,0)' % (i, i, i)


    c = str(len(df1)+2)
    sheet['G1'] = '=SUM(G3:G' + c +')'

    blue_fill = PatternFill(start_color="7cb5eb", end_color="7cb5eb", fill_type="solid")
    rule2 = Rule(type="expression", dxf=DifferentialStyle(fill=blue_fill))
    rule2.formula = ['AND(G3<200,G3<>0)']
    sheet.conditional_formatting.add(('G3:G' + c), rule2)

    pink_fill = PatternFill(start_color="FFC0CB", end_color="FFC0CB", fill_type="solid")
    rule2 = Rule(type="expression", dxf=DifferentialStyle(fill=pink_fill))
    rule2.formula = ['OR(R3<=0,R3="нет значения")']
    sheet.conditional_formatting.add(('R3:R' + c), rule2)

    rule3 = Rule(type="expression", dxf=DifferentialStyle(fill=pink_fill))
    rule3.formula = ['OR(Y3="Да", Z3="Да")']
    sheet.conditional_formatting.add(('B3:B' + c), rule3)


    wb.save(file_path + 'Заказ ' + prk_number + ' ' + now1 + '.xlsx')

    add_message('Готово!')


        





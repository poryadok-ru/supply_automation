import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
from time import time
import glob
import os
from python_calamine.pandas import pandas_monkeypatch

pandas_monkeypatch()

pd.options.mode.chained_assignment = None


os.chdir(os.path.dirname(os.path.abspath(__file__)))


def process_and_group_excel(excel_file):
    """
    Обрабатывает Excel файл, перенося комментарии из одного листа в другой, добавляя строки.
    Подсвечивает комментарии прошлого месяца желтым цветом.
    Группирует строки по кодам продукта и подсвечивает строки с позициями, у которых есть хотя бы один комментарий за прошлый месяц.

    Args:
        excel_file (str): Имя Excel файла.
    """
    t = time()
    try:
        # 1. Чтение данных из Excel файла с помощью pandas
        df_comments = pd.read_excel(excel_file, sheet_name="общее", engine="calamine")
        df_products = pd.read_excel(excel_file, sheet_name="по кодам", engine="calamine")

        # 2. Фильтрация комментариев (удаление NaN и "-")
        df_comments = df_comments[df_comments["Комментарий"].notna()]
        df_comments = df_comments[df_comments["Комментарий"].str.strip() != "-"]

        # 3. Преобразование столбца "Месяц и год" в datetime
        df_comments['Месяц и год'] = pd.to_datetime(df_comments['Месяц и год'], errors='coerce')

        # Удаляем строки с NaT значениями
        df_comments = df_comments[df_comments['Месяц и год'].notna()]

        # 4. Создание словаря "код продукта - комментарии"
        product_comments = df_comments.groupby("Код продукта").apply(lambda x: list(zip(x["Комментарий"], x["Месяц и год"]))).to_dict()

        # 5. Загрузка рабочей книги с помощью openpyxl
        wb = load_workbook(excel_file)
        sheet_products = wb["по кодам"]

        # Определение прошлого месяца
        today = datetime.now()
        last_month = today.replace(day=1) - timedelta(days=1)
        last_month_str = last_month.strftime("%m")  # Получаем только месяц

        # 6. Добавление комментариев на лист с продуктами
        row_offset = 0  # Смещение для вставки строк
        original_rows = list(sheet_products.iter_rows(min_row=2, values_only=True))  # Сохраняем исходные строки

        for row_idx, row_data in enumerate(original_rows, start=2):
            if row_data:  # Проверяем, что строка не пустая
                product_code = row_data[0]  # Предполагаем, что код продукта в первом столбце

                if product_code in product_comments:
                    comments = product_comments[product_code]

                    # Находим номер строки для вставки (с учетом смещения).  Вставляем *после* текущей строки.
                    insert_row_index = row_idx + row_offset + 1  # +1 для вставки после строки с product_code

                    # Вставляем комментарии сразу после строки с кодом продукта
                    for comment, date in comments:
                        # Вставляем строку со сдвигом
                        sheet_products.insert_rows(insert_row_index)

                        # Записываем комментарий в столбец "Номенклатура"
                        sheet_products.cell(row=insert_row_index, column=2).value = comment

                        # Подсвечиваем строку, если дата комментария относится к прошлому месяцу
                        month_in_date = date.strftime("%m")  # Извлекаем месяц из datetime

                        if month_in_date == last_month_str:
                            for col in range(1, sheet_products.max_column + 1):
                                sheet_products.cell(row=insert_row_index, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                        # Увеличиваем смещение и индекс для вставки
                        row_offset += 1
                        insert_row_index += 1

        # 7. Группировка строк по кодам продукта
        max_row = sheet_products.max_row

        # Iterate through the rows, identifying product code rows and grouping subsequent rows
        group_start = 1  # Start from the first row
        i = 1
        while i <= max_row:
            # Check if the first cell in the row contains a product code
            if sheet_products.cell(row=i, column=1).value is not None and str(sheet_products.cell(row=i, column=1).value).strip() != "":
                # Found a product code row, group subsequent rows until the next product code
                group_end = i + 1
                while group_end <= max_row:
                    if sheet_products.cell(row=group_end, column=1).value is not None and str(sheet_products.cell(row=group_end, column=1).value).strip() != "":
                        # Found the next product code, group the rows before it
                        break
                    group_end += 1

                # Create the row group (excluding the product code row itself)
                if i + 1 < group_end:
                    sheet_products.row_dimensions.group(start=i + 1, end=group_end - 1, hidden=False)  # Changed to False to keep rows visible

                # Check if any of the comments are highlighted in yellow
                has_yellow_comment = False
                for row in range(i + 1, group_end):
                    for col in range(1, sheet_products.max_column + 1):
                        cell = sheet_products.cell(row=row, column=col)
                        if cell.fill and cell.fill.start_color.rgb == "00FFFF00":  # Check for yellow fill
                            has_yellow_comment = True
                            break
                    if has_yellow_comment:
                        break

                if has_yellow_comment:
                    # Color the product code row
                    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    for col in range(1, sheet_products.max_column + 1):
                        sheet_products.cell(row=i, column=col).fill = fill

                # Move to the next product code
                i = group_end
            else:
                i += 1

        # 8. Сохранение изменений
        wb.save("updated_" + excel_file)
        print(f"Комментарии добавлены и строки сгруппированы в файл 'updated_{excel_file}'")

    except FileNotFoundError:
        print(f"Ошибка: Файл '{excel_file}' не найден.")
    except KeyError as e:
        print(f"Ошибка: Лист с именем '{e}' не найден в Excel файле.")
    except Exception as e:
        print(f"Произошла ошибка: {e}")
    print(time() - t)



import os
import glob
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from time import time


def npsview(file_path, add_message):
    if not file_path.endswith(os.sep):
        file_path = file_path + os.sep
    t = time()
    add_message('Стартуем!')

    files = list(glob.iglob(file_path + "Исходники\\*.xlsx"))
    today = datetime.now()
    last_month = today.replace(day=1) - timedelta(days=1)
    last_month_str = last_month.strftime("%m")

    for excel_file in files:
        print(excel_file)
        try:
            # === Шаг 1. Прочитать нужные листы ===
            dfs = pd.read_excel(excel_file, sheet_name=["1", "2", "по кодам"], engine="calamine")

            # Объединяем листы "1" и "2" в общий датафрейм комментариев
            df_comments_nach = pd.concat([dfs["1"], dfs["2"]], ignore_index=True)
            df_comments = df_comments_nach.copy()
            df_products = dfs["по кодам"]

            # === Шаг 2. Подготовка комментариев ===
            df_comments = df_comments[df_comments["Комментарий"].notna()]
            df_comments = df_comments[df_comments["Комментарий"].str.strip() != "-"]

            # Корректное преобразование дат
            # Иногда Excel сохраняет как число (например, 45123), иногда как текст — обрабатываем оба случая
            def safe_to_datetime(x):
                try:
                    if pd.isna(x):
                        return pd.NaT
                    # Если это число (Excel serial date)
                    if isinstance(x, (int, float)):
                        return pd.to_datetime("1899-12-30") + pd.to_timedelta(int(x), "D")
                    # Если это строка — пробуем разные форматы
                    return pd.to_datetime(x, errors="coerce", dayfirst=True)
                except Exception:
                    return pd.NaT

            df_comments["Месяц и год"] = df_comments["Месяц и год"].apply(safe_to_datetime)
            df_comments = df_comments[df_comments["Месяц и год"].notna()]

            product_comments = df_comments.groupby("Код продукта").apply(
                lambda x: list(zip(x["Комментарий"], x["Месяц и год"]))
            ).to_dict()

            # === Шаг 3. Формируем список строк с комментами и метками ===
            rows_to_write = []
            is_code_row = []     # метка: строка-код продукта, нужна для outline
            is_yellow_row = []   # метка: строка выделяется жёлтым

            for idx, row in df_products.iterrows():
                # product row
                rows_to_write.append(list(row) + [None])    # +1 столбец для заливки
                is_code_row.append(True)
                is_yellow_row.append(False)                 # жёлтым станет позже, если надо

                code = row.iloc[0]
                supplier = row.iloc[2] if len(row) > 2 else None
                manager  = row.iloc[3] if len(row) > 3 else None
                has_yellow_comment = False

                comments_data = product_comments.get(code, [])
                comment_rows = []
                yellow_flags = []
                for comment, cdate in comments_data:
                    # Подсвечиваем, если комментарий сделан в прошлом месяце текущего года
                    is_yellow = (cdate.year == last_month.year) and (cdate.month == last_month.month)
                    comment_row = [None] * len(row)   # все NaN
                    comment_row[1] = comment
                    if len(row) > 2: comment_row[2] = supplier
                    if len(row) > 3: comment_row[3] = manager
                    comment_row.append(is_yellow)
                    comment_rows.append(comment_row)
                    yellow_flags.append(is_yellow)
                    if is_yellow:
                        has_yellow_comment = True

                # Добавить все комментарии под текущим продуктом
                rows_to_write.extend(comment_rows)
                is_code_row.extend([False] * len(comment_rows))
                is_yellow_row.extend(yellow_flags)

                # Если хоть один из комментариев жёлтый, саму код-строку тоже надо красить
                if has_yellow_comment:
                    is_yellow_row[-(len(comment_rows) + 1)] = True

            # === Шаг 4. Записываем в новый Excel ===
            out_df = pd.DataFrame(rows_to_write)
            out_df.columns = list(df_products.columns) + ["_yellow"]
            output_file = file_path + 'upd_' + os.path.basename(excel_file)
            with pd.ExcelWriter(output_file, engine="openpyxl", mode="w") as writer:
                out_df.drop(columns="_yellow").to_excel(writer, sheet_name="по кодам", index=False)
                # df_comments_nach.to_excel(writer, sheet_name="общее", index=False)

            # === Шаг 5. Раскраска и outline ===
            wb = load_workbook(output_file)
            sh = wb["по кодам"]
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            # Массив меток смещаем на header (он строка 1)
            is_code_row2 = [0] + is_code_row
            is_yellow_row2 = [False] + is_yellow_row

            max_row = sh.max_row
            max_col = sh.max_column

            # Закрашиваем строки
            for i in range(2, max_row+1):
                if is_yellow_row2[i-1]:
                    for col in range(1, max_col+1):
                        sh.cell(row=i, column=col).fill = yellow_fill

            # Группировка outline
            i = 2
            while i <= max_row:
                if is_code_row2[i-1]:   # если это строка кода продукта
                    group_start = i+1
                    # ищем, где закончились комментарии под этим кодом
                    group_end = group_start
                    while group_end <= max_row and not is_code_row2[group_end-1]:
                        group_end += 1
                    # если комментариев >0, то группируем
                    if group_end > group_start:
                        sh.row_dimensions.group(start=group_start, end=group_end-1, hidden=False)
                    i = group_end
                else:
                    i += 1

            wb.save(output_file)
            print(f"Комментарии добавлены и строки сгруппированы в файл '{excel_file}'")

        except FileNotFoundError:
            print(f"Ошибка: Файл '{excel_file}' не найден.")
        except KeyError as e:
            print(f"Ошибка: Лист с именем '{e}' не найден в Excel файле.")
        except Exception as e:
            print(f"Произошла ошибка: {e}")
        print(time() - t)
        add_message('Готов ' + os.path.basename(excel_file))

    add_message(f'Готово за {time() - t}')